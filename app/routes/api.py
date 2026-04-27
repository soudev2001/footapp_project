
# =====================
# ADMIN SEED ENDPOINTS
# =====================

import jwt
import os
import datetime
import smtplib
from flask import Blueprint, jsonify, request, current_app, render_template, url_for, make_response
from app.models import serialize_doc, serialize_docs
from app.services import (
    get_player_service, get_club_service, get_event_service,
    get_match_service, get_post_service, get_user_service,
    get_team_service, get_notification_service, get_contract_service,
    get_shop_service, get_project_service, get_parent_link_service,
    get_subscription_service, get_isy_service,
    get_analytics_service, get_member_onboarding_service, get_billing_service,
    get_announcement_service,
    get_platform_management_service, get_platform_analytics_service,
    get_parent_monitoring_service, get_fan_engagement_service, get_media_service,
)
from app.services.messaging_service import MessagingService
from app.services.db import mongo
from functools import wraps
from bson import ObjectId
from email.mime.text import MIMEText

api_bp = Blueprint('api', __name__, url_prefix='/api')

ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'svg', 'webp'}


# ============================================================
# JWT HELPERS
# ============================================================

def token_required(f):
    """Decorator that requires a valid JWT Bearer token."""
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization', None)
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'message': 'Token manquant'}), 401
        token = auth_header.split(' ')[1]
        try:
            payload = jwt.decode(
                token,
                current_app.config['JWT_SECRET_KEY'],
                algorithms=['HS256']
            )
            if payload.get('type') == 'refresh':
                return jsonify({'success': False, 'message': 'Cannot use refresh token for API calls'}), 401
            request.current_user = payload
        except jwt.ExpiredSignatureError:
            return jsonify({'success': False, 'message': 'Token has expired!'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'success': False, 'message': 'Token is invalid!'}), 401

        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    """Decorator that requires specific user roles."""
    def decorator(f):
        @wraps(f)
        @token_required
        def decorated(*args, **kwargs):
            user_role = request.current_user.get('role', '')
            if user_role == 'admin' or user_role in roles:
                return f(*args, **kwargs)
            return jsonify({'success': False, 'message': 'Insufficient permissions'}), 403
        return decorated
    return decorator


def _get_club_doc(club_id):
    if not club_id or not ObjectId.is_valid(str(club_id)):
        return None
    return mongo.db.clubs.find_one({'_id': ObjectId(club_id)})


def _get_club_personalization(club):
    if not club:
        return {}

    personalization = dict(club.get('personalization', {}))
    colors = club.get('colors', {})

    personalization['clubName'] = club.get('name', '')
    
    # Only include logoUrl if it's a valid URL (starts with / or http)
    logo = club.get('logo')
    if logo and isinstance(logo, str):
        logo = logo.strip()
    if logo and (logo.startswith('/') or logo.startswith('http')):
        personalization['logoUrl'] = logo
    else:
        personalization.pop('logoUrl', None)
    
    # Only include coverUrl if it's a valid URL
    cover = club.get('cover_url')
    if cover and isinstance(cover, str):
        cover = cover.strip()
    if cover and (cover.startswith('/') or cover.startswith('http')):
        personalization['coverUrl'] = cover
    else:
        personalization.pop('coverUrl', None)
    
    personalization.setdefault('primaryColor', colors.get('primary', '#22c55e'))
    personalization.setdefault('accentColor', colors.get('secondary', '#16a34a'))

    return personalization


def _build_auth_user(user):
    profile = user.get('profile', {})
    club = _get_club_doc(user.get('club_id')) if user.get('club_id') else None

    return {
        'id': str(user['_id']),
        'email': user['email'],
        'role': user.get('role', 'fan'),
        'club_id': str(user.get('club_id', '')) if user.get('club_id') else None,
        'profile': {
            'first_name': profile.get('first_name', ''),
            'last_name': profile.get('last_name', ''),
            'avatar': profile.get('avatar', ''),
            'phone': profile.get('phone', ''),
        },
        'account_status': user.get('account_status', 'active'),
        'club_personalization': _get_club_personalization(club),
    }


def _save_club_asset(club_id, file_storage, asset_name):
    ext = file_storage.filename.rsplit('.', 1)[-1].lower() if '.' in file_storage.filename else ''
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return None, 'Type de fichier invalide'

    upload_dir = os.path.join(current_app.static_folder, 'uploads', 'clubs', str(club_id))
    os.makedirs(upload_dir, exist_ok=True)

    filename = f'{asset_name}.{ext}'
    file_storage.save(os.path.join(upload_dir, filename))
    return f'/static/uploads/clubs/{club_id}/{filename}', None

@api_bp.route('/admin/seed-players', methods=['POST'])
@role_required('admin')
def api_seed_players():
    """Seed 18 joueurs démo pour le club de l'admin connecté"""
    from app.services.seed_data import seed_18_players
    club_id = request.current_user.get('club_id')
    team_id = request.json.get('team_id') if request.is_json else None
    try:
        players = seed_18_players(club_id, team_id)
        return jsonify({'success': True, 'message': f'{len(players)} joueurs créés', 'count': len(players)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@api_bp.route('/admin/seed-all', methods=['POST'])
@role_required('admin')
def api_seed_all():
    """Seed complet de la base (données démo)"""
    from app.services.seed_data import seed_all
    try:
        seed_all()
        return jsonify({'success': True, 'message': 'Seed complet effectué'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/admin/seed-coach-data', methods=['POST'])
@role_required('admin')
def api_seed_coach_data():
    """Seed des données coach (tactiques, lineups, drills, plans) pour le club de l'admin connecté"""
    from app.services.seed_data import seed_coach_data
    club_id = request.current_user.get('club_id')
    team_id = request.json.get('team_id') if request.is_json else None
    try:
        seed_coach_data(club_id, team_id)
        return jsonify({'success': True, 'message': 'Seed coach data effectué'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def transform_player_for_frontend(player):
    """Transform player document to include profile structure expected by frontend."""
    doc = serialize_doc(player)
    
    # Prioritize individual name fields if they exist
    first_name = player.get('first_name')
    last_name = player.get('last_name')
    
    # Fallback to splitting 'name' field if components are missing
    if not first_name and not last_name:
        name_parts = (player.get('name') or '').split(' ', 1)
        first_name = name_parts[0] if name_parts else ''
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        
    doc['profile'] = {
        'first_name': first_name or '',
        'last_name': last_name or '',
        'avatar': player.get('photo', player.get('avatar', ''))
    }
    return doc


def transform_players_for_frontend(players):
    """Transform list of players for frontend."""
    return [transform_player_for_frontend(p) for p in players]


# ============================================================
# JWT HELPERS
# ============================================================

def generate_token(user, expires_hours=None):
    """Generate a signed JWT token for a user."""
    if expires_hours is None:
        expires_hours = current_app.config.get('JWT_EXPIRATION_HOURS', 24)
    payload = {
        'user_id': str(user['_id']),
        'email': user.get('email', ''),
        'role': user.get('role', 'fan'),
        'club_id': str(user.get('club_id', '')) if user.get('club_id') else None,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=expires_hours),
        'iat': datetime.datetime.utcnow(),
        'type': 'access'
    }
    return jwt.encode(payload, current_app.config['JWT_SECRET_KEY'], algorithm='HS256')


def generate_refresh_token(user):
    """Generate a longer-lived refresh token."""
    days = current_app.config.get('JWT_REFRESH_EXPIRATION_DAYS', 30)
    payload = {
        'user_id': str(user['_id']),
        'exp': datetime.datetime.utcnow() + datetime.timedelta(days=days),
        'iat': datetime.datetime.utcnow(),
        'type': 'refresh'
    }
    return jwt.encode(payload, current_app.config['JWT_SECRET_KEY'], algorithm='HS256')


def token_required(f):
    """Decorator that requires a valid JWT Bearer token."""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            if auth_header.startswith('Bearer '):
                token = auth_header.split(" ", 1)[1]

        if not token:
            return jsonify({'success': False, 'message': 'Token is missing!'}), 401

        try:
            payload = jwt.decode(
                token,
                current_app.config['JWT_SECRET_KEY'],
                algorithms=['HS256']
            )
            if payload.get('type') == 'refresh':
                return jsonify({'success': False, 'message': 'Cannot use refresh token for API calls'}), 401
            request.current_user = payload
        except jwt.ExpiredSignatureError:
            return jsonify({'success': False, 'message': 'Token has expired!'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'success': False, 'message': 'Token is invalid!'}), 401

        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    """Decorator that requires specific user roles."""
    def decorator(f):
        @wraps(f)
        @token_required
        def decorated(*args, **kwargs):
            user_role = request.current_user.get('role', '')
            if user_role == 'admin' or user_role in roles:
                return f(*args, **kwargs)
            return jsonify({'success': False, 'message': 'Insufficient permissions'}), 403
        return decorated
    return decorator


# ============================================================
# AUTH ENDPOINTS
# ============================================================

@api_bp.route('/auth/login', methods=['POST'])
def api_login():
    """Authenticate user and return JWT tokens."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Request body required'}), 400

    email = (data.get('email') or '').strip().lower()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({'success': False, 'error': 'Email and password required'}), 400

    user_service = get_user_service()
    user = user_service.verify_password(email, password)

    if not user:
        return jsonify({'success': False, 'error': 'Invalid credentials'}), 401

    access_token = generate_token(user)
    refresh_token = generate_refresh_token(user)

    return jsonify({
        'success': True,
        'access_token': access_token,
        'refresh_token': refresh_token,
        'user': _build_auth_user(user)
    })


@api_bp.route('/auth/register', methods=['POST'])
def api_register():
    """Register a new user and return JWT tokens."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Request body required'}), 400

    email = (data.get('email') or '').strip().lower()
    password = data.get('password', '')
    role = data.get('role', 'fan')
    club_id = data.get('club_id')

    if not email or not password:
        return jsonify({'success': False, 'error': 'Email and password required'}), 400

    if len(password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters'}), 400

    user_service = get_user_service()
    existing = user_service.get_by_email(email)
    if existing:
        return jsonify({'success': False, 'error': 'Email already registered'}), 409

    profile = {
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
        'phone': data.get('phone', ''),
    }

    user_id = user_service.create(email, password, role=role, club_id=club_id, profile=profile)
    user = user_service.get_by_id(user_id)

    access_token = generate_token(user)
    refresh_token = generate_refresh_token(user)

    return jsonify({
        'success': True,
        'access_token': access_token,
        'refresh_token': refresh_token,
        'user': {
            'id': str(user['_id']),
            'email': user['email'],
            'role': user.get('role', 'fan'),
            'club_id': str(user.get('club_id', '')) if user.get('club_id') else None,
        }
    }), 201


@api_bp.route('/auth/refresh', methods=['POST'])
def api_refresh():
    """Refresh access token using a valid refresh token."""
    data = request.get_json()
    refresh = data.get('refresh_token') if data else None
    if not refresh:
        return jsonify({'success': False, 'error': 'Refresh token required'}), 400

    try:
        payload = jwt.decode(refresh, current_app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        if payload.get('type') != 'refresh':
            return jsonify({'success': False, 'error': 'Invalid token type'}), 401
    except jwt.ExpiredSignatureError:
        return jsonify({'success': False, 'error': 'Refresh token expired, please login again'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'success': False, 'error': 'Invalid refresh token'}), 401

    user_service = get_user_service()
    user = user_service.get_by_id(payload['user_id'])
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 404

    access_token = generate_token(user)
    return jsonify({
        'success': True,
        'access_token': access_token
    })


@api_bp.route('/auth/me', methods=['GET'])
@token_required
def api_me():
    """Get current authenticated user profile."""
    user_service = get_user_service()
    user = user_service.get_by_id(request.current_user['user_id'])
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 404

    player_service = get_player_service()
    player = player_service.get_by_user(str(user['_id']))

    result = _build_auth_user(user)

    if player:
        result['player'] = serialize_doc(player)

    return jsonify({'success': True, 'data': result})


# ============================================================
# PLAYER ENDPOINTS
# ============================================================

@api_bp.route('/player/profile', methods=['GET'])
@token_required
def get_player_profile():
    """Get current player's full profile."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player profile not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(player)})


@api_bp.route('/player/profile', methods=['PUT'])
@token_required
def update_player_profile():
    """Update current user profile."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400

    user_service = get_user_service()
    user_service.update_profile(request.current_user['user_id'], data)
    return jsonify({'success': True, 'message': 'Profile updated'})


@api_bp.route('/player/stats', methods=['GET'])
@token_required
def get_player_stats():
    """Get current player's stats."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404

    return jsonify({
        'success': True,
        'data': {
            'stats': player.get('stats', {}),
            'technical_ratings': player.get('technical_ratings', {}),
            'physical_records': player.get('physical_records', []),
        }
    })


@api_bp.route('/player/contracts', methods=['GET'])
@token_required
def get_player_contracts():
    """Get current player's contracts."""
    contract_service = get_contract_service()
    contracts = contract_service.get_by_user(request.current_user['user_id'])
    return jsonify({
        'success': True,
        'data': serialize_docs(contracts)
    })


@api_bp.route('/player/contracts/<contract_id>/respond', methods=['POST'])
@token_required
def respond_contract(contract_id):
    """Accept or reject a contract offer."""
    data = request.get_json()
    action = data.get('action') if data else None
    if action not in ('active', 'rejected'):
        return jsonify({'success': False, 'error': 'action must be active or rejected'}), 400
    contract_service = get_contract_service()
    contract_service.respond_to_offer(contract_id, action)
    return jsonify({'success': True, 'message': f'Contract {action}'})


# ============================================================
# PLAYERS API (public / coach)
# ============================================================

@api_bp.route('/players', methods=['GET'])
def get_players():
    """Get all players (optionally filtered by club_id query param)."""
    club_id = request.args.get('club_id')
    team_id = request.args.get('team_id')
    player_service = get_player_service()

    if club_id:
        players = player_service.get_by_club(club_id, team_id=team_id)
    else:
        players = player_service.get_all()

    return jsonify({
        'success': True,
        'count': len(players),
        'data': transform_players_for_frontend(players)
    })


@api_bp.route('/players/<player_id>', methods=['GET'])
def get_player(player_id):
    """Get single player by ID."""
    player_service = get_player_service()
    player = player_service.get_by_id(player_id)
    if player:
        return jsonify({'success': True, 'data': serialize_doc(player)})
    return jsonify({'success': False, 'error': 'Player not found'}), 404


# ============================================================
# CLUBS API
# ============================================================

@api_bp.route('/clubs', methods=['GET'])
def get_clubs():
    """Get all clubs."""
    club_service = get_club_service()
    clubs = club_service.get_all()
    return jsonify({
        'success': True,
        'count': len(clubs),
        'data': serialize_docs(clubs)
    })


@api_bp.route('/clubs/<club_id>', methods=['GET'])
def get_club(club_id):
    """Get single club by ID."""
    club_service = get_club_service()
    club = club_service.get_by_id(club_id)
    if club:
        return jsonify({'success': True, 'data': serialize_doc(club)})
    return jsonify({'success': False, 'error': 'Club not found'}), 404


@api_bp.route('/clubs/<club_id>/stats', methods=['GET'])
def get_club_stats(club_id):
    """Get club statistics."""
    club_service = get_club_service()
    stats = club_service.get_stats(club_id)
    return jsonify({'success': True, 'data': stats})


@api_bp.route('/clubs/<club_id>/players', methods=['GET'])
def get_club_players(club_id):
    """Get all players of a club."""
    team_id = request.args.get('team_id')
    player_service = get_player_service()
    players = player_service.get_by_club(club_id, team_id=team_id)
    return jsonify({
        'success': True,
        'count': len(players),
        'data': serialize_docs(players)
    })


# ============================================================
# TEAM ENDPOINTS
# ============================================================

@api_bp.route('/teams', methods=['GET'])
@token_required
def get_teams():
    """Get teams for current user's club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        # Fallback: lookup user's club_id from database
        user_service = get_user_service()
        user = user_service.get_by_id(request.current_user.get('user_id'))
        club_id = str(user.get('club_id')) if user and user.get('club_id') else None
    if not club_id:
        return jsonify({'success': True, 'data': []})
    team_service = get_team_service()
    teams = team_service.get_by_club(club_id)
    return jsonify({'success': True, 'data': serialize_docs(teams)})


@api_bp.route('/teams/<team_id>', methods=['GET'])
@token_required
def get_team(team_id):
    """Get team details."""
    team_service = get_team_service()
    team = team_service.get_by_id(team_id)
    if not team:
        return jsonify({'success': False, 'error': 'Team not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(team)})


@api_bp.route('/teams/<team_id>/players', methods=['GET'])
@token_required
def get_team_players(team_id):
    """Get players of a specific team."""
    player_service = get_player_service()
    club_id = request.current_user.get('club_id')
    # Try team_id filter first; fall back to all club players if empty
    players = player_service.get_by_club(club_id, team_id=team_id) if club_id else []
    if not players:
        team_service = get_team_service()
        players = team_service.get_players(team_id)
    return jsonify({
        'success': True,
        'count': len(players),
        'data': transform_players_for_frontend(players)
    })


# ============================================================
# EVENTS / CALENDAR API
# ============================================================

@api_bp.route('/events', methods=['GET'])
def get_events():
    """Get events. Filter by club_id, team_id, type query params."""
    club_id = request.args.get('club_id')
    event_type = request.args.get('type')
    event_service = get_event_service()

    if club_id and event_type:
        events = event_service.get_by_type(club_id, event_type)
    elif club_id:
        events = event_service.get_by_club(club_id)
    else:
        events = event_service.get_all()

    return jsonify({
        'success': True,
        'count': len(events),
        'data': serialize_docs(events)
    })


@api_bp.route('/events/<event_id>', methods=['GET'])
def get_event(event_id):
    """Get single event."""
    event_service = get_event_service()
    event = event_service.get_by_id(event_id)
    if event:
        return jsonify({'success': True, 'data': serialize_doc(event)})
    return jsonify({'success': False, 'error': 'Event not found'}), 404


@api_bp.route('/calendar/upcoming', methods=['GET'])
@token_required
def get_calendar_upcoming():
    """Get upcoming events + matches for current user's club."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    limit = request.args.get('limit', 20, type=int)

    if not club_id:
        return jsonify({'success': True, 'data': {'events': [], 'matches': []}})

    event_service = get_event_service()
    match_service = get_match_service()

    events = event_service.get_upcoming(club_id, team_id=team_id, limit=limit)
    matches = match_service.get_upcoming(club_id, team_id=team_id, limit=limit)

    return jsonify({
        'success': True,
        'data': {
            'events': serialize_docs(events),
            'matches': serialize_docs(matches),
        }
    })


@api_bp.route('/events/<event_id>/rsvp', methods=['POST'])
@token_required
def update_rsvp(event_id):
    """RSVP to an event (present/absent/uncertain)."""
    data = request.get_json()
    status = data.get('status') if data else None
    if status not in ('present', 'absent', 'uncertain'):
        return jsonify({'success': False, 'error': 'Status must be present, absent, or uncertain'}), 400

    event_service = get_event_service()
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    player_id = str(player['_id']) if player else request.current_user['user_id']

    event_service.set_attendance(event_id, player_id, status)
    return jsonify({'success': True, 'message': f'RSVP updated to {status}'})


@api_bp.route('/events/<event_id>/attendance', methods=['GET'])
@token_required
def get_event_attendance(event_id):
    """Get attendance list for an event."""
    event_service = get_event_service()
    attendance = event_service.get_attendance_list(event_id)
    return jsonify({'success': True, 'data': serialize_docs(attendance)})


# ============================================================
# MATCHES API
# ============================================================

@api_bp.route('/matches', methods=['GET'])
def get_matches():
    """Get matches. Filter by club_id query param."""
    club_id = request.args.get('club_id')
    match_service = get_match_service()

    if club_id:
        matches = match_service.get_by_club(club_id)
    else:
        matches = match_service.get_all()

    return jsonify({
        'success': True,
        'count': len(matches),
        'data': serialize_docs(matches)
    })


@api_bp.route('/matches/upcoming', methods=['GET'])
@token_required
def get_upcoming_matches():
    """Get upcoming matches for current user's club."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    limit = request.args.get('limit', 10, type=int)

    if not club_id:
        return jsonify({'success': True, 'data': []})

    match_service = get_match_service()
    matches = match_service.get_upcoming(club_id, team_id=team_id, limit=limit)
    return jsonify({'success': True, 'data': serialize_docs(matches)})


@api_bp.route('/matches/results', methods=['GET'])
@token_required
def get_match_results():
    """Get completed matches."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    limit = request.args.get('limit', 10, type=int)

    if not club_id:
        return jsonify({'success': True, 'data': []})

    match_service = get_match_service()
    matches = match_service.get_completed(club_id, team_id=team_id, limit=limit)
    return jsonify({'success': True, 'data': serialize_docs(matches)})


@api_bp.route('/matches/<match_id>', methods=['GET'])
def get_match(match_id):
    """Get single match details."""
    match_service = get_match_service()
    match = match_service.get_by_id(match_id)
    if not match:
        return jsonify({'success': False, 'error': 'Match not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(match)})


@api_bp.route('/matches/<match_id>/lineup', methods=['GET'])
def get_match_lineup(match_id):
    """Get match lineup."""
    match_service = get_match_service()
    lineup = match_service.get_lineup(match_id)
    return jsonify({'success': True, 'data': serialize_docs(lineup) if lineup else []})


@api_bp.route('/matches/season-stats', methods=['GET'])
@token_required
def get_season_stats():
    """Get season stats for current club."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})
    match_service = get_match_service()
    stats = match_service.get_season_stats(club_id, team_id=team_id)
    return jsonify({'success': True, 'data': stats})


# ============================================================
# POSTS / FEED API
# ============================================================

@api_bp.route('/posts', methods=['GET'])
def get_posts():
    """Get posts. Filter by club_id query param."""
    club_id = request.args.get('club_id')
    category = request.args.get('category')
    limit = request.args.get('limit', 20, type=int)
    post_service = get_post_service()

    if club_id and category:
        posts = post_service.get_by_category(club_id, category, limit=limit)
    elif club_id:
        posts = post_service.get_by_club(club_id, limit=limit)
    else:
        posts = post_service.get_all(limit=limit)

    return jsonify({
        'success': True,
        'count': len(posts),
        'data': serialize_docs(posts)
    })


@api_bp.route('/posts/<post_id>', methods=['GET'])
def get_post(post_id):
    """Get single post."""
    post_service = get_post_service()
    post = post_service.get_by_id(post_id)
    if post:
        return jsonify({'success': True, 'data': serialize_doc(post)})
    return jsonify({'success': False, 'error': 'Post not found'}), 404


@api_bp.route('/posts/<post_id>/like', methods=['POST'])
@token_required
def like_post(post_id):
    """Like a post."""
    post_service = get_post_service()
    post_service.like(post_id)
    return jsonify({'success': True, 'message': 'Post liked'})


@api_bp.route('/posts/<post_id>/comment', methods=['POST'])
@token_required
def comment_post(post_id):
    """Add a comment to a post."""
    data = request.get_json()
    text = data.get('text') if data else None
    if not text:
        return jsonify({'success': False, 'error': 'Comment text required'}), 400
    post_service = get_post_service()
    post_service.add_comment(post_id, request.current_user['user_id'], text)
    return jsonify({'success': True, 'message': 'Comment added'})


@api_bp.route('/posts/search', methods=['GET'])
@token_required
def search_posts():
    """Search posts."""
    club_id = request.current_user.get('club_id')
    query = request.args.get('q', '')
    if not club_id or not query:
        return jsonify({'success': True, 'data': []})
    post_service = get_post_service()
    posts = post_service.search(club_id, query)
    return jsonify({'success': True, 'data': serialize_docs(posts)})


# ============================================================
# MESSAGING API
# ============================================================

@api_bp.route('/messages/conversations', methods=['GET'])
@token_required
def get_conversations():
    """Get conversation previews for current user."""
    club_id = request.current_user.get('club_id')
    user_id = request.current_user['user_id']
    svc = MessagingService(mongo.db)
    previews = svc.get_last_messages_preview(user_id, club_id)

    # Convert dict previews to a list with proper id/name/type for the frontend
    conversations = []
    for key, preview in previews.items():
        parts = key.split('_', 1)
        conv_type = parts[0]  # direct, team, channel
        ref_id = parts[1] if len(parts) > 1 else key
        conv = {
            '_id': key,
            'id': key,
            'type': conv_type,
            'last_message': preview.get('content', ''),
            'last_message_at': str(preview.get('time', '')),
            'unread_count': 0,
        }
        if conv_type == 'direct':
            conv['other_user_id'] = ref_id
            other_user = mongo.db.users.find_one({'_id': ObjectId(ref_id)})
            if other_user:
                profile = other_user.get('profile', {})
                conv['name'] = f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip() or other_user.get('email', 'Utilisateur')
                conv['avatar'] = profile.get('avatar', '')
            else:
                conv['name'] = 'Utilisateur'
        elif conv_type == 'team':
            conv['team_id'] = ref_id
            team = mongo.db.teams.find_one({'_id': ObjectId(ref_id)})
            conv['name'] = team.get('name', 'Équipe') if team else 'Équipe'
        else:
            conv['name'] = ref_id
        conversations.append(conv)

    conversations.sort(key=lambda c: c.get('last_message_at', ''), reverse=True)
    return jsonify({'success': True, 'data': conversations})


@api_bp.route('/messages/direct/<other_user_id>', methods=['GET'])
@token_required
def get_direct_messages(other_user_id):
    """Get direct message history with another user."""
    limit = request.args.get('limit', 50, type=int)
    svc = MessagingService(mongo.db)
    messages = svc.get_direct_messages(request.current_user['user_id'], other_user_id, limit=limit)
    return jsonify({'success': True, 'data': serialize_docs(messages)})


@api_bp.route('/messages/team/<team_id>', methods=['GET'])
@token_required
def get_team_messages(team_id):
    """Get team chat messages."""
    limit = request.args.get('limit', 50, type=int)
    svc = MessagingService(mongo.db)
    messages = svc.get_team_messages(team_id, limit=limit)
    return jsonify({'success': True, 'data': serialize_docs(messages)})


@api_bp.route('/messages/channel/<channel_id>', methods=['GET'])
@token_required
def get_channel_messages(channel_id):
    """Get channel messages."""
    limit = request.args.get('limit', 50, type=int)
    svc = MessagingService(mongo.db)
    messages = svc.get_channel_messages(channel_id, limit=limit)
    return jsonify({'success': True, 'data': serialize_docs(messages)})


@api_bp.route('/messages/send', methods=['POST'])
@token_required
def send_message():
    """Send a message (direct, team, or channel)."""
    data = request.get_json()
    if not data or not data.get('content'):
        return jsonify({'success': False, 'error': 'Content required'}), 400

    svc = MessagingService(mongo.db)
    msg_id = svc.send_message(
        sender_id=request.current_user['user_id'],
        content=data['content'],
        receiver_id=data.get('receiver_id'),
        team_id=data.get('team_id'),
        msg_type=data.get('type', 'direct'),
        channel_id=data.get('channel_id'),
    )
    return jsonify({'success': True, 'message_id': str(msg_id)}), 201


@api_bp.route('/messages/<message_id>/read', methods=['POST'])
@token_required
def mark_message_read(message_id):
    """Mark a message as read."""
    svc = MessagingService(mongo.db)
    svc.mark_as_read(message_id, request.current_user['user_id'])
    return jsonify({'success': True})


@api_bp.route('/messages/unread-count', methods=['GET'])
@token_required
def get_unread_count():
    """Get unread messages count."""
    svc = MessagingService(mongo.db)
    count = svc.get_unread_count(request.current_user['user_id'])
    return jsonify({'success': True, 'count': count})


@api_bp.route('/messages/channels', methods=['GET'])
@token_required
def get_channels():
    """Get channels for current user's club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    svc = MessagingService(mongo.db)
    channels = svc.get_channels(club_id, user_id=request.current_user['user_id'])
    return jsonify({'success': True, 'data': serialize_docs(channels)})


# ============================================================
# NOTIFICATIONS API
# ============================================================

@api_bp.route('/notifications', methods=['GET'])
@token_required
def get_notifications():
    """Get notifications for current user."""
    user_id = request.current_user['user_id']
    raw = list(mongo.db.notifications.find(
        {'$or': [{'user_id': user_id}, {'user_id': ObjectId(user_id)}]}
    ).sort('sent_at', -1).limit(50))
    # Ensure created_at exists for frontend compatibility
    for n in raw:
        if 'created_at' not in n and 'sent_at' in n:
            n['created_at'] = n['sent_at']
    return jsonify({'success': True, 'data': serialize_docs(raw)})


@api_bp.route('/notifications/mark-read/<notification_id>', methods=['POST'])
@token_required
def mark_notification_read(notification_id):
    """Mark a notification as read."""
    mongo.db.notifications.update_one(
        {'_id': ObjectId(notification_id)},
        {'$set': {'read': True}}
    )
    return jsonify({'success': True})


@api_bp.route('/notifications/mark-all-read', methods=['POST'])
@token_required
def mark_all_notifications_read():
    """Mark all notifications as read for current user."""
    user_id = request.current_user['user_id']
    mongo.db.notifications.update_many(
        {'$or': [{'user_id': user_id}, {'user_id': ObjectId(user_id)}], 'read': False},
        {'$set': {'read': True}}
    )
    return jsonify({'success': True})


@api_bp.route('/notifications/register-device', methods=['POST'])
@token_required
def register_device_token():
    """Register a device push notification token (FCM/Expo)."""
    data = request.get_json()
    push_token = data.get('push_token') or data.get('fcmToken') if data else None
    if not push_token:
        return jsonify({'success': False, 'error': 'push_token required'}), 400

    platform = data.get('platform', 'expo')
    user_id = request.current_user['user_id']

    mongo.db.device_tokens.update_one(
        {'user_id': user_id, 'token': push_token},
        {'$set': {
            'user_id': user_id,
            'token': push_token,
            'platform': platform,
            'updated_at': datetime.datetime.utcnow()
        }},
        upsert=True
    )
    return jsonify({'success': True, 'message': 'Device token registered'})


# ============================================================
# COACH ENDPOINTS
# ============================================================

@api_bp.route('/coach/dashboard', methods=['GET'])
@role_required('coach')
def coach_dashboard():
    """Get coach dashboard data."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')

    if not club_id:
        return jsonify({'success': True, 'data': {}})

    player_service = get_player_service()
    match_service = get_match_service()
    event_service = get_event_service()

    players = player_service.get_by_club(club_id, team_id=team_id)
    upcoming_matches = match_service.get_upcoming(club_id, team_id=team_id, limit=5)
    upcoming_events = event_service.get_upcoming(club_id, team_id=team_id, limit=5)
    season_stats = match_service.get_season_stats(club_id, team_id=team_id)
    top_scorers = player_service.get_top_scorers(club_id, limit=5)

    injured = [p for p in players if p.get('status') == 'injured']

    return jsonify({
        'success': True,
        'data': {
            'total_players': len(players),
            'injured_players': serialize_docs(injured),
            'upcoming_matches': serialize_docs(upcoming_matches),
            'upcoming_events': serialize_docs(upcoming_events),
            'season_stats': season_stats,
            'top_scorers': serialize_docs(top_scorers),
        }
    })


@api_bp.route('/coach/roster', methods=['GET'])
@role_required('coach')
def coach_roster():
    """Get full roster for coach."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    player_service = get_player_service()
    players = player_service.get_by_club(club_id, team_id=team_id)

    return jsonify({'success': True, 'data': transform_players_for_frontend(players)})


@api_bp.route('/coach/convocation', methods=['POST'])
@role_required('coach')
def send_convocation():
    """Send convocation for a match/event with tactical instructions."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400

    event_id = data.get('event_id')
    player_ids = data.get('player_ids', [])

    if not event_id or not player_ids:
        return jsonify({'success': False, 'error': 'event_id and player_ids required'}), 400

    club_id = request.current_user.get('club_id')
    notification_service = get_notification_service()
    event_service = get_event_service()
    player_service = get_player_service()
    event = event_service.get_by_id(event_id)

    event_title = event.get("title", "Événement") if event else "Événement"
    event_date = event.get("date", "") if event else ""
    event_location = event.get("location", "") if event else ""

    # Save convocation document
    convocation_id = player_service.save_convocation(club_id, event_id, {
        'formation': data.get('formation', '4-3-3'),
        'starters': data.get('starters', []),
        'substitutes': data.get('substitutes', []),
        'captains': data.get('captains', []),
        'set_pieces': data.get('set_pieces', {}),
        'player_instructions': data.get('player_instructions', {}),
        'message': data.get('message', ''),
        'match_date': data.get('match_date'),
        'player_ids': player_ids,
    })

    # Per-player: notification + email
    starters_list = data.get('starters', [])
    player_instructions = data.get('player_instructions', {})
    set_pieces = data.get('set_pieces', {})
    formation = data.get('formation', '4-3-3')

    SET_PIECE_LABELS = {
        'penalties': 'Tireur de pénaltys',
        'free_kicks_direct': 'Coups francs directs',
        'free_kicks_indirect': 'Coups francs indirects',
        'corners_left': 'Corners gauche',
        'corners_right': 'Corners droit',
    }

    for pid in player_ids:
        # Find player's position in the lineup
        position = None
        slot_key = None
        if isinstance(starters_list, list):
            for i, sid in enumerate(starters_list):
                if sid == pid:
                    position = f"Titulaire (poste {i + 1})"
                    slot_key = str(i)
                    break
        if not position and pid in data.get('substitutes', []):
            position = "Remplaçant"

        # Get player's role instructions
        role_info = None
        if slot_key and slot_key in player_instructions:
            role_info = player_instructions[slot_key]
        # Also check by player ID key
        if not role_info and pid in player_instructions:
            role_info = player_instructions[pid]
        # Check all keys (slot keys like "ST-0")
        for k, v in player_instructions.items():
            if isinstance(starters_list, list):
                # Find which slot this player is in
                for skey, sdata in (data.get('slot_map', {}) or {}).items():
                    if sdata == pid and k == skey:
                        role_info = v
                        break

        # Set piece duties for this player
        sp_duties = []
        for sp_key, sp_ids in set_pieces.items():
            if pid in sp_ids:
                label = SET_PIECE_LABELS.get(sp_key, sp_key)
                idx = sp_ids.index(pid) + 1
                sp_duties.append(f"{label} (priorité {idx})")

        # Resolve the player's linked user_id for notification + email
        player = player_service.get_by_id(pid)
        player_user_id = str(player.get('user_id', '')) if player and player.get('user_id') else None
        notify_user_id = player_user_id or pid

        # Create notification with link to match prep (target the USER, not the player doc)
        notification_service.create_notification(
            user_id=notify_user_id,
            title='Convocation',
            message=f'Vous êtes convoqué pour: {event_title}',
            type='convocation',
            link=f'/player/match-prep/{convocation_id}'
        )
        event_service.set_attendance(event_id, pid, 'convoked')

        # Send email — look up user doc for email
        try:
            if player:
                email = None
                player_name = player.get('name', 'Joueur')
                if player_user_id:
                    from app.services import get_user_service
                    user_svc = get_user_service()
                    user_doc = user_svc.get_by_id(player_user_id)
                    if user_doc:
                        email = user_doc.get('email')
                        profile = user_doc.get('profile', {})
                        player_name = f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip() or player_name
                if email:
                    from app.services.email_service import send_convocation_email
                    send_convocation_email(
                        to_email=email,
                        player_name=player_name,
                        event_info={'title': event_title, 'date': str(event_date), 'location': event_location},
                        position=position,
                        role_info=role_info,
                        set_piece_duties=sp_duties if sp_duties else None,
                    )
        except Exception as e:
            current_app.logger.error(f"Erreur envoi email convocation à {pid}: {e}")

    return jsonify({'success': True, 'message': f'{len(player_ids)} players convoked', 'convocation_id': convocation_id})


@api_bp.route('/coach/lineup', methods=['GET'])
@role_required('coach')
def get_lineup():
    """Get active lineup."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    if not club_id:
        return jsonify({'success': True, 'data': None})
    player_service = get_player_service()
    lineup = player_service.get_active_lineup(club_id, team_id)
    return jsonify({'success': True, 'data': serialize_doc(lineup) if lineup else None})


@api_bp.route('/coach/lineup', methods=['POST'])
@role_required('coach')
def save_lineup():
    """Save a lineup."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400

    club_id = request.current_user.get('club_id')
    player_service = get_player_service()
    result = player_service.save_lineup(
        club_id=club_id,
        formation=data.get('formation', '4-3-3'),
        starters=data.get('starters', []),
        team_id=data.get('team_id'),
        substitutes=data.get('substitutes', []),
        name=data.get('name', ''),
        captains=data.get('captains', []),
        set_pieces=data.get('set_pieces', {}),
        player_instructions=data.get('player_instructions', {})
    )
    return jsonify({'success': True, 'data': str(result)})


@api_bp.route('/coach/tactics', methods=['GET'])
@role_required('coach')
def get_tactics():
    """Get saved tactic presets."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    player_service = get_player_service()
    tactics = player_service.get_tactic_presets(club_id, team_id)
    return jsonify({'success': True, 'data': serialize_docs(tactics)})


@api_bp.route('/coach/tactics', methods=['POST'])
@role_required('coach')
def save_tactic():
    """Save a tactic (create or update)."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    club_id = request.current_user.get('club_id')
    player_service = get_player_service()
    try:
        preset_id = player_service.save_tactic_preset(
            club_id=club_id,
            team_id=data.get('team_id'),
            name=data.get('name', 'Sans nom'),
            description=data.get('description', ''),
            formation=data.get('formation', '4-3-3'),
            starters=data.get('starters', []),
            substitutes=data.get('substitutes', []),
            instructions={
                'passing_style': data.get('passing_style'),
                'pressing': data.get('pressing'),
                'defensive_block': data.get('defensive_block'),
                'marking': data.get('marking'),
                'tempo': data.get('tempo'),
                'width': data.get('width'),
                'play_space': data.get('play_space'),
                'gk_distribution': data.get('gk_distribution'),
                'counter_pressing': data.get('counter_pressing', False),
                'mentality': data.get('mentality', 'balanced'),
                'defensive_shape': data.get('defensive_shape', 'normal'),
                'buildup_style': data.get('buildup_style', 'mixed'),
                'transition_speed': data.get('transition_speed', 'balanced'),
                'offside_trap': data.get('offside_trap', False),
                'creative_freedom': data.get('creative_freedom', 'balanced'),
                'defensive_width': data.get('defensive_width', 'normal'),
                'pressing_trigger': data.get('pressing_trigger', 'opponent_half'),
            },
            captains=data.get('captains', []),
            set_pieces=data.get('set_pieces', {}),
            player_instructions=data.get('player_instructions', {}),
            tactic_id=data.get('id'),
        )
        return jsonify({'success': True, 'data': str(preset_id)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/coach/tactics/<tactic_id>', methods=['DELETE'])
@role_required('coach')
def delete_tactic(tactic_id):
    """Delete a tactic."""
    player_service = get_player_service()
    player_service.delete_tactic_preset(tactic_id)
    return jsonify({'success': True, 'message': 'Tactique supprimée'})


@api_bp.route('/coach/events', methods=['GET'])
@role_required('coach')
def coach_get_events():
    """Get coach events list."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    event_service = get_event_service()
    events = event_service.get_by_club(club_id)
    return jsonify({'success': True, 'data': serialize_docs(events)})


@api_bp.route('/coach/attendance', methods=['GET'])
@role_required('coach')
def coach_get_attendance():
    """Get attendance data for an event."""
    event_id = request.args.get('event_id')
    if not event_id:
        return jsonify({'success': True, 'data': {}})
    event_service = get_event_service()
    attendance = event_service.get_attendance(event_id)
    return jsonify({'success': True, 'data': attendance})


@api_bp.route('/coach/convocation', methods=['GET'])
@role_required('coach')
def get_convocation():
    """Get convocation data (upcoming matches + roster)."""
    club_id = request.current_user.get('club_id')
    team_id = request.args.get('team_id')
    if not club_id:
        return jsonify({'success': True, 'data': {'matches': [], 'players': []}})
    match_service = get_match_service()
    player_service = get_player_service()
    matches = match_service.get_upcoming(club_id, team_id=team_id, limit=10)
    players = player_service.get_by_club(club_id, team_id=team_id)
    return jsonify({'success': True, 'data': {
        'matches': serialize_docs(matches),
        'players': serialize_docs(players),
    }})


# ============================================================
# DATABASE UTILITIES (Development only)
# ============================================================

@api_bp.route('/seed', methods=['POST'])
def seed_database():
    """Seed database with demo data."""
    from app.services.seed_data import seed_all
    try:
        seed_all()
        return jsonify({'success': True, 'message': 'Database seeded successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/stats', methods=['GET'])
def get_db_stats():
    """Get database statistics."""
    from app.services.db import get_stats
    stats = get_stats()
    return jsonify({'success': True, 'data': stats})


@api_bp.route('/health', methods=['GET'])
def health_check():
    """API health check."""
    return jsonify({
        'success': True,
        'status': 'healthy',
        'version': '2.0.0',
        'app': 'FootLogic V2'
    })


# ============================================================
# SWAGGER DOCS
# ============================================================

@api_bp.route('/docs')
def swagger_ui():
    """Serve Swagger UI for API Documentation."""
    return render_template('api/docs.html')


# ============================================================
# PLAYER EXTENDED ENDPOINTS
# ============================================================

@api_bp.route('/player/documents', methods=['GET'])
@token_required
def get_player_documents():
    """Get player documents (license, medical cert, ID)."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    docs = player.get('documents', {})
    return jsonify({'success': True, 'data': docs})


ALLOWED_DOC_TYPES = {'licence', 'medical_cert', 'id_card', 'insurance', 'photo'}
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png'}


@api_bp.route('/player/documents/<doc_type>', methods=['POST'])
@token_required
def upload_player_document(doc_type):
    """Upload a player document file."""
    if doc_type not in ALLOWED_DOC_TYPES:
        return jsonify({'success': False, 'error': 'Type de document invalide'}), 400

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier fourni'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': 'Nom de fichier vide'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'success': False, 'error': 'Extension non autorisée'}), 400

    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404

    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'player_docs',
                              str(player['_id']))
    os.makedirs(upload_dir, exist_ok=True)

    filename = f"{doc_type}.{ext}"
    filepath = os.path.join(upload_dir, filename)
    file.save(filepath)

    rel_url = f"/static/uploads/player_docs/{player['_id']}/{filename}"
    player_service.update_documents(str(player['_id']), doc_type, 'provided', rel_url)

    return jsonify({'success': True, 'data': {'url': rel_url, 'doc_type': doc_type}})


@api_bp.route('/player/evolution', methods=['GET'])
@token_required
def get_player_evolution():
    """Get player evolution data (ratings, physical history, evaluations)."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    return jsonify({
        'success': True,
        'data': {
            'technical_ratings': player.get('technical_ratings', {}),
            'physical_history': player.get('physical_history', []),
            'evaluations': player.get('evaluations', []),
            'stats': player.get('stats', {}),
        }
    })


# ============================================================
# PLAYER: ENHANCED ENDPOINTS
# ============================================================

@api_bp.route('/player/dashboard/stats', methods=['GET'])
@token_required
def player_dashboard_stats():
    """Get player performance dashboard."""
    from app.services import get_player_analytics_service
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    svc = get_player_analytics_service()
    data = svc.get_player_dashboard(str(player['_id']))
    return jsonify({'success': True, 'data': data})


@api_bp.route('/player/dashboard/rankings', methods=['GET'])
@token_required
def player_dashboard_rankings():
    """Get player position rankings within team."""
    from app.services import get_player_analytics_service
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    team_id = player.get('team_id')
    if not team_id:
        return jsonify({'success': True, 'data': []})
    svc = get_player_analytics_service()
    data = svc.get_team_rankings(str(team_id))
    return jsonify({'success': True, 'data': data})


@api_bp.route('/player/goals', methods=['GET'])
@token_required
def player_get_goals():
    """Get player personal goals."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    goals = list(mongo.db.player_goals.find({'player_id': player['_id']}).sort('created_at', -1))
    return jsonify({'success': True, 'data': serialize_docs(goals)})


@api_bp.route('/player/goals', methods=['POST'])
@token_required
def player_create_goal():
    """Create a personal goal."""
    data = request.get_json()
    if not data or not data.get('title'):
        return jsonify({'success': False, 'error': 'Titre requis'}), 400
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    doc = {
        'player_id': player['_id'],
        'category': data.get('category', 'technical'),
        'title': data['title'],
        'description': data.get('description', ''),
        'target_value': data.get('target_value', 0),
        'current_value': data.get('current_value', 0),
        'target_date': data.get('target_date'),
        'status': 'active',
        'created_at': datetime.datetime.utcnow(),
    }
    result = mongo.db.player_goals.insert_one(doc)
    return jsonify({'success': True, 'goal_id': str(result.inserted_id)}), 201


@api_bp.route('/player/goals/<goal_id>', methods=['PUT'])
@token_required
def player_update_goal(goal_id):
    """Update a personal goal."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data'}), 400
    updates = {}
    for field in ['title', 'description', 'target_value', 'current_value', 'target_date', 'status', 'category']:
        if field in data:
            updates[field] = data[field]
    if updates:
        updates['updated_at'] = datetime.datetime.utcnow()
        mongo.db.player_goals.update_one({'_id': ObjectId(goal_id)}, {'$set': updates})
    return jsonify({'success': True, 'message': 'Objectif mis à jour'})


@api_bp.route('/player/goals/<goal_id>', methods=['DELETE'])
@token_required
def player_delete_goal(goal_id):
    """Delete a personal goal."""
    mongo.db.player_goals.delete_one({'_id': ObjectId(goal_id)})
    return jsonify({'success': True, 'message': 'Objectif supprimé'})


@api_bp.route('/player/training/schedule', methods=['GET'])
@token_required
def player_training_schedule():
    """Get player training schedule."""
    from app.services import get_training_service
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    team_id = player.get('team_id')
    if not team_id:
        return jsonify({'success': True, 'data': []})
    svc = get_training_service()
    plans = svc.get_plans(str(team_id))
    return jsonify({'success': True, 'data': plans})


@api_bp.route('/player/training/drills', methods=['GET'])
@token_required
def player_training_drills():
    """Get drills available to player."""
    from app.services import get_training_service
    svc = get_training_service()
    drills = svc.get_drills({})
    return jsonify({'success': True, 'data': drills})


@api_bp.route('/player/training/drills/<drill_id>/complete', methods=['POST'])
@token_required
def player_complete_drill(drill_id):
    """Mark a drill as completed by the player."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    data = request.get_json() or {}
    doc = {
        'player_id': player['_id'],
        'drill_id': drill_id,
        'completed_at': datetime.datetime.utcnow(),
        'note': data.get('note', ''),
    }
    mongo.db.drill_completions.update_one(
        {'player_id': player['_id'], 'drill_id': drill_id},
        {'$set': doc},
        upsert=True,
    )
    return jsonify({'success': True, 'message': 'Exercice marqué complété'})


@api_bp.route('/player/training/notes', methods=['GET'])
@token_required
def player_training_notes_get():
    """Get personal training notes."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    notes = list(mongo.db.player_training_notes.find(
        {'player_id': player['_id']}
    ).sort('created_at', -1).limit(50))
    return jsonify({'success': True, 'data': serialize_docs(notes)})


@api_bp.route('/player/training/notes', methods=['POST'])
@token_required
def player_training_notes_post():
    """Create a personal training note."""
    data = request.get_json()
    if not data or not data.get('content'):
        return jsonify({'success': False, 'error': 'content requis'}), 400
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    doc = {
        'player_id': player['_id'],
        'content': data['content'],
        'exercise_name': data.get('exercise_name', ''),
        'created_at': datetime.datetime.utcnow(),
    }
    result = mongo.db.player_training_notes.insert_one(doc)
    return jsonify({'success': True, 'note_id': str(result.inserted_id)}), 201


@api_bp.route('/player/match-prep/<convocation_id>', methods=['GET'])
@token_required
def player_match_prep(convocation_id):
    """Get match preparation data for a player. Lineup is time-gated (24h before match)."""
    player_service = get_player_service()
    convocation = player_service.get_convocation(convocation_id)

    if not convocation:
        return jsonify({'success': False, 'error': 'Convocation not found'}), 404

    user_id = request.current_user.get('user_id') or str(request.current_user.get('_id', ''))
    player_id = request.current_user.get('player_id') or user_id

    # Check if lineup should be visible (24h before match_date)
    lineup_visible = False
    match_date = convocation.get('match_date')
    if match_date:
        if isinstance(match_date, str):
            try:
                from datetime import datetime
                match_date = datetime.fromisoformat(match_date)
            except (ValueError, TypeError):
                match_date = None
        if match_date:
            from datetime import datetime, timedelta
            now = datetime.utcnow()
            if now >= match_date - timedelta(hours=24):
                lineup_visible = True

    # Find this player's slot info
    my_slot = None
    my_instructions = None
    my_set_pieces = []
    starters = convocation.get('starters', [])
    player_instructions = convocation.get('player_instructions', {})
    set_pieces = convocation.get('set_pieces', {})

    # Check if player is in starters
    if isinstance(starters, list):
        for i, sid in enumerate(starters):
            if str(sid) == str(player_id):
                my_slot = {'index': i, 'type': 'starter'}
                break
    # Check subs
    if not my_slot:
        for sid in convocation.get('substitutes', []):
            if str(sid) == str(player_id):
                my_slot = {'type': 'substitute'}
                break

    # Get player instructions (try multiple key formats)
    for key, val in player_instructions.items():
        if isinstance(starters, list):
            try:
                idx = int(key.split('-')[-1]) if '-' in key else int(key)
                if idx < len(starters) and str(starters[idx]) == str(player_id):
                    my_instructions = val
                    break
            except (ValueError, IndexError):
                pass
        if key == str(player_id):
            my_instructions = val
            break

    # Set piece duties
    SP_LABELS = {
        'penalties': 'Pénaltys', 'free_kicks_direct': 'Coups francs directs',
        'free_kicks_indirect': 'Coups francs indirects',
        'corners_left': 'Corners gauche', 'corners_right': 'Corners droit',
    }
    for sp_key, sp_ids in set_pieces.items():
        if str(player_id) in [str(x) for x in sp_ids]:
            idx = [str(x) for x in sp_ids].index(str(player_id))
            my_set_pieces.append({'key': sp_key, 'label': SP_LABELS.get(sp_key, sp_key), 'priority': idx + 1})

    # Build response
    response = {
        'id': convocation.get('id'),
        'event_id': convocation.get('event_id'),
        'formation': convocation.get('formation'),
        'match_date': str(convocation.get('match_date', '')),
        'message': convocation.get('message', ''),
        'sent_at': str(convocation.get('sent_at', '')),
        'captains': convocation.get('captains', []),
        'my_slot': my_slot,
        'my_instructions': my_instructions,
        'my_set_pieces': my_set_pieces,
        'lineup_visible': lineup_visible,
    }

    # Only include full lineup if time-gated check passes
    if lineup_visible:
        response['starters'] = [str(s) if s else None for s in starters]
        response['substitutes'] = [str(s) for s in convocation.get('substitutes', [])]
        response['set_pieces'] = {k: [str(x) for x in v] for k, v in set_pieces.items()}
        response['player_instructions'] = player_instructions

    # Get event info
    event_id = convocation.get('event_id')
    if event_id:
        event_service = get_event_service()
        event = event_service.get_by_id(event_id)
        if event:
            response['event'] = {
                'title': event.get('title', ''),
                'date': str(event.get('date', '')),
                'location': event.get('location', ''),
                'type': event.get('type', event.get('event_type', '')),
            }

    return jsonify({'success': True, 'data': response})


@api_bp.route('/player/convocations', methods=['GET'])
@token_required
def player_convocations():
    """List all convocations received by the current player."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    player_id = str(player['_id'])

    convocations_raw = list(mongo.db.convocations.find({
        '$or': [
            {'starters': {'$in': [player_id, ObjectId(player_id) if len(player_id) == 24 else player_id]}},
            {'substitutes': {'$in': [player_id, ObjectId(player_id) if len(player_id) == 24 else player_id]}},
        ]
    }).sort('match_date', -1).limit(30))

    result = []
    event_service = get_event_service()
    for conv in convocations_raw:
        conv_id = str(conv['_id'])
        event_info = {}
        if conv.get('event_id'):
            try:
                ev = event_service.get_by_id(str(conv['event_id']))
                if ev:
                    event_info = {
                        'title': ev.get('title', ''),
                        'date': str(ev.get('date', '')),
                        'location': ev.get('location', ''),
                        'type': ev.get('type', ev.get('event_type', '')),
                    }
            except Exception:
                pass
        starters = [str(s) for s in conv.get('starters', [])]
        substitutes = [str(s) for s in conv.get('substitutes', [])]
        slot_type = 'starter' if player_id in starters else ('substitute' if player_id in substitutes else 'unknown')
        result.append({
            'id': conv_id,
            'match_date': str(conv.get('match_date', '')),
            'opponent': conv.get('opponent', ''),
            'location': conv.get('location', conv.get('venue', '')),
            'slot_type': slot_type,
            'formation': conv.get('formation', ''),
            'event': event_info,
            'created_at': str(conv.get('created_at', '')),
        })
    return jsonify({'success': True, 'data': result})


@api_bp.route('/player/matches', methods=['GET'])
@token_required
def player_matches():
    """List matches for the player's team with personal stats per match."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    player_id = str(player['_id'])
    team_id = player.get('team_id')

    query = {'team_id': team_id} if team_id else {}
    matches_raw = list(mongo.db.matches.find(query).sort('date', -1).limit(30))

    result = []
    for m in matches_raw:
        match_id = str(m['_id'])
        # Find personal stats for this player in this match
        personal = {}
        player_stats = m.get('player_stats', [])
        if isinstance(player_stats, list):
            for ps in player_stats:
                if str(ps.get('player_id', '')) == player_id:
                    personal = ps
                    break
        elif isinstance(player_stats, dict):
            personal = player_stats.get(player_id, {})

        result.append({
            'id': match_id,
            'date': str(m.get('date', '')),
            'opponent': m.get('opponent', m.get('away_team', '')),
            'home_score': m.get('home_score', m.get('score_home')),
            'away_score': m.get('away_score', m.get('score_away')),
            'result': m.get('result', ''),
            'status': m.get('status', 'scheduled'),
            'location': m.get('location', m.get('venue', '')),
            'competition': m.get('competition', ''),
            'personal_stats': {
                'goals': personal.get('goals', 0),
                'assists': personal.get('assists', 0),
                'rating': personal.get('rating'),
                'minutes_played': personal.get('minutes_played', personal.get('minutes', 0)),
                'yellow_cards': personal.get('yellow_cards', 0),
                'red_cards': personal.get('red_cards', 0),
            },
        })
    return jsonify({'success': True, 'data': result})


@api_bp.route('/player/tactics/current', methods=['GET'])
@token_required
def player_current_tactic():
    """Return the active/latest tactic for the player's team (read-only)."""
    player_service = get_player_service()
    player = player_service.get_by_user(request.current_user['user_id'])
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    team_id = player.get('team_id')
    club_id = player.get('club_id')

    query = {}
    if team_id:
        query['team_id'] = team_id
    elif club_id:
        query['club_id'] = club_id

    tactic = None
    if query:
        tactic = mongo.db.tactics.find_one(query, sort=[('updated_at', -1)])
    if not tactic:
        return jsonify({'success': True, 'data': None})

    t = dict(tactic)
    t['id'] = str(t.pop('_id'))
    for k in ('team_id', 'club_id', 'created_by'):
        if k in t:
            t[k] = str(t[k])
    return jsonify({'success': True, 'data': t})


# ============================================================
# COACH EXTENDED ENDPOINTS
# ============================================================

@api_bp.route('/coach/players', methods=['POST'])
@role_required('coach')
def coach_add_player():
    """Add a new player to the club."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400

    club_id = request.current_user.get('club_id')
    user_service = get_user_service()
    player_service = get_player_service()

    email = (data.get('email') or '').strip().lower()
    if email:
        existing = user_service.get_by_email(email)
        if existing:
            return jsonify({'success': False, 'error': 'Email already registered'}), 409

    profile = {
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
        'phone': data.get('phone', ''),
    }

    password = data.get('password', 'changeme123')
    user_id = user_service.create(email or f"player_{datetime.datetime.utcnow().timestamp()}@footapp.local",
                                   password, role='player', club_id=club_id, profile=profile)

    first_name = data.get('first_name', '')
    last_name = data.get('last_name', '')
    name = f"{first_name} {last_name}".strip()

    player_id = player_service.create(
        user_id=str(user_id),
        club_id=club_id,
        name=name,
        first_name=first_name,
        last_name=last_name,
        team_id=data.get('team_id'),
        jersey_number=data.get('jersey_number'),
        position=data.get('position', 'MID'),
        birth_date=data.get('birth_date'),
        photo=data.get('photo'),
    )
    return jsonify({'success': True, 'player_id': str(player_id)}), 201


@api_bp.route('/coach/players/<player_id>', methods=['PUT'])
@role_required('coach')
def coach_edit_player(player_id):
    """Edit player details."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    player_service = get_player_service()
    player_service.update(player_id, data)
    return jsonify({'success': True, 'message': 'Player updated'})


@api_bp.route('/coach/players/<player_id>', methods=['DELETE'])
@role_required('coach')
def coach_delete_player(player_id):
    """Remove player from roster."""
    player_service = get_player_service()
    player_service.delete(player_id)
    return jsonify({'success': True, 'message': 'Player removed'})


@api_bp.route('/coach/players/<player_id>/ratings', methods=['POST'])
@role_required('coach')
def coach_update_ratings(player_id):
    """Update player technical ratings."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    player_service = get_player_service()
    player_service.update_technical_ratings(player_id, data)
    return jsonify({'success': True, 'message': 'Ratings updated'})


@api_bp.route('/coach/players/<player_id>/evaluation', methods=['POST'])
@role_required('coach')
def coach_add_evaluation(player_id):
    """Add coach evaluation for a player."""
    data = request.get_json()
    if not data or not data.get('comment'):
        return jsonify({'success': False, 'error': 'Comment required'}), 400
    player_service = get_player_service()
    player_service.add_evaluation(player_id, {
        'coach_id': request.current_user['user_id'],
        'comment': data['comment'],
        'rating': data.get('rating'),
        'date': datetime.datetime.utcnow().isoformat(),
    })
    return jsonify({'success': True, 'message': 'Evaluation added'})


@api_bp.route('/coach/players/<player_id>/physical', methods=['POST'])
@role_required('coach')
def coach_add_physical(player_id):
    """Record physical metrics for a player."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    player_service = get_player_service()
    record = {
        'weight': data.get('weight'),
        'height': data.get('height'),
        'vma': data.get('vma'),
        'date': datetime.datetime.utcnow().isoformat(),
    }
    player_service.add_physical_record(player_id, record)
    return jsonify({'success': True, 'message': 'Physical record added'})


@api_bp.route('/coach/attendance/update', methods=['POST'])
@role_required('coach')
def coach_update_attendance():
    """Bulk update attendance for an event."""
    data = request.get_json()
    if not data or not data.get('event_id') or not data.get('attendance'):
        return jsonify({'success': False, 'error': 'event_id and attendance required'}), 400

    event_service = get_event_service()
    for item in data['attendance']:
        event_service.set_attendance(data['event_id'], item['player_id'], item['status'])
    return jsonify({'success': True, 'message': 'Attendance updated'})


@api_bp.route('/coach/events', methods=['POST'])
@role_required('coach')
def coach_create_event():
    """Create a new event (training, match, meeting)."""
    data = request.get_json()
    if not data or not data.get('title'):
        return jsonify({'success': False, 'error': 'Title required'}), 400

    club_id = request.current_user.get('club_id')
    event_service = get_event_service()
    event_id = event_service.create(
        club_id=club_id,
        team_id=data.get('team_id'),
        title=data['title'],
        event_type=data.get('type', 'training'),
        date=data.get('date'),
        end_date=data.get('end_date'),
        location=data.get('location', ''),
        description=data.get('description', ''),
    )
    return jsonify({'success': True, 'event_id': str(event_id['_id'])}), 201


@api_bp.route('/coach/events/<event_id>', methods=['PUT'])
@role_required('coach')
def coach_edit_event(event_id):
    """Edit an existing event."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    event_service = get_event_service()
    event_service.update(event_id, data)
    return jsonify({'success': True, 'message': 'Event updated'})


@api_bp.route('/coach/events/<event_id>', methods=['DELETE'])
@role_required('coach')
def coach_delete_event(event_id):
    """Delete an event."""
    event_service = get_event_service()
    event_service.delete(event_id)
    return jsonify({'success': True, 'message': 'Event deleted'})


@api_bp.route('/coach/matches/<match_id>/score', methods=['POST'])
@role_required('coach')
def coach_update_score(match_id):
    """Update match score."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    match_service = get_match_service()
    match_service.set_score(match_id, data.get('home', 0), data.get('away', 0),
                            status=data.get('status', 'in_progress'))
    return jsonify({'success': True, 'message': 'Score updated'})


@api_bp.route('/coach/matches/<match_id>/event', methods=['POST'])
@role_required('coach')
def coach_add_match_event(match_id):
    """Add match event (goal, card, substitution)."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    match_service = get_match_service()
    match_service.add_event(match_id, data)
    return jsonify({'success': True, 'message': 'Match event added'})


@api_bp.route('/coach/matches', methods=['POST'])
@role_required('coach')
def coach_create_match():
    """Create a new match."""
    data = request.get_json()
    if not data or not data.get('opponent'):
        return jsonify({'success': False, 'error': 'Opponent required'}), 400
    club_id = request.current_user.get('club_id')
    match_service = get_match_service()
    match_id = match_service.create(
        club_id=club_id,
        opponent=data['opponent'],
        date=data.get('date'),
        is_home=data.get('is_home', True),
        team_id=data.get('team_id'),
        competition=data.get('competition', ''),
        location=data.get('location', ''),
    )
    return jsonify({'success': True, 'match_id': str(match_id)}), 201


# ============================================================
# ADMIN ENDPOINTS
# ============================================================

@api_bp.route('/admin/dashboard', methods=['GET'])
@role_required('admin')
def admin_dashboard():
    """Get admin dashboard data."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})

    user_service = get_user_service()
    team_service = get_team_service()
    player_service = get_player_service()
    event_service = get_event_service()
    match_service = get_match_service()

    members = user_service.get_members_by_club(club_id)
    teams = team_service.get_by_club(club_id)
    players = player_service.get_by_club(club_id)
    upcoming_events = event_service.get_upcoming(club_id, limit=5)
    club_service = get_club_service()
    club = club_service.get_by_id(club_id)

    role_counts = {}
    for m in members:
        r = m.get('role', 'fan')
        role_counts[r] = role_counts.get(r, 0) + 1

    pending = [m for m in members if m.get('account_status') == 'pending']

    return jsonify({
        'success': True,
        'data': {
            'club': serialize_doc(club) if club else {},
            'total_members': len(members),
            'total_teams': len(teams),
            'total_players': len(players),
            'role_counts': role_counts,
            'pending_invitations': len(pending),
            'upcoming_events': serialize_docs(upcoming_events),
            'teams': serialize_docs(teams),
        }
    })


@api_bp.route('/admin/members', methods=['GET'])
@role_required('admin')
def admin_members():
    """Get all members of the club with enriched player data."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    role = request.args.get('role')
    status = request.args.get('status')
    user_service = get_user_service()
    members = user_service.get_members_by_club(club_id)
    if role:
        members = [m for m in members if m.get('role') == role]
    if status:
        members = [m for m in members if m.get('account_status') == status]

    # Enrich with player data
    player_service = get_player_service()
    user_ids = [m['_id'] for m in members]
    players_list = list(player_service.collection.find({'user_id': {'$in': user_ids}}))
    players_map = {str(p['user_id']): p for p in players_list}

    for member in members:
        player = players_map.get(str(member['_id']))
        if player:
            # Format birth_date for JSON
            birth_date = player.get('birth_date')
            if birth_date:
                birth_date = birth_date.isoformat() if hasattr(birth_date, 'isoformat') else str(birth_date)
            member['player_data'] = {
                'player_id': str(player['_id']),
                'team_id': str(player.get('team_id', '')) if player.get('team_id') else '',
                'jersey_number': player.get('jersey_number'),
                'position': player.get('position'),
                'photo': player.get('photo'),
                'birth_date': birth_date,
                'height': player.get('height'),
                'weight': player.get('weight'),
                'documents': player.get('documents', {}),
                'license_number': player.get('license_number'),
                'status': player.get('status', 'active')
            }

    return jsonify({'success': True, 'count': len(members), 'data': serialize_docs(members)})


@api_bp.route('/admin/members', methods=['POST'])
@role_required('admin')
def admin_add_member():
    """Add a new member to the club."""
    data = request.get_json()
    if not data or not data.get('email'):
        return jsonify({'success': False, 'error': 'Email required'}), 400

    club_id = request.current_user.get('club_id')
    user_service = get_user_service()
    email = data['email'].strip().lower()

    existing = user_service.get_by_email(email)
    if existing:
        return jsonify({'success': False, 'error': 'Email already registered'}), 409

    profile = {
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
        'phone': data.get('phone', ''),
    }
    user_id = user_service.create(
        email, data.get('password', 'changeme123'),
        role=data.get('role', 'player'), club_id=club_id, profile=profile
    )
    return jsonify({'success': True, 'user_id': str(user_id)}), 201


@api_bp.route('/admin/invite', methods=['POST'])
@role_required('admin')
def admin_invite_member():
    """Invite a new member by email with invitation token."""
    import secrets
    data = request.get_json()
    if not data or not data.get('email'):
        return jsonify({'success': False, 'error': 'Email required'}), 400

    club_id = request.current_user.get('club_id')
    user_service = get_user_service()
    notification_service = get_notification_service()
    email = data['email'].strip().lower()

    existing = user_service.get_by_email(email)
    if existing:
        if existing.get('account_status') == 'active':
            return jsonify({'success': False, 'error': 'Cet utilisateur est déjà actif'}), 409
        # Resend invitation
        if not existing.get('invitation_token'):
            token = secrets.token_urlsafe(32)
            user_service.collection.update_one(
                {'_id': existing['_id']},
                {'$set': {'invitation_token': token}}
            )
            existing['invitation_token'] = token
        notification_service.send_invitation(existing)
        return jsonify({'success': True, 'message': 'Invitation renvoyée'})

    # Create new pending user with invitation token
    profile = {
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
    }
    user = user_service.create_pending_user(
        email=email,
        role=data.get('role', 'player'),
        club_id=club_id,
        profile=profile,
    )
    notification_service.send_invitation(user)
    return jsonify({'success': True, 'message': 'Invitation envoyée', 'user_id': str(user['_id'])}), 201


@api_bp.route('/admin/members/<user_id>', methods=['PUT'])
@role_required('admin')
def admin_edit_member(user_id):
    """Edit a member with full player data support."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    user_service = get_user_service()
    from bson import ObjectId

    update = {}
    # Profile fields
    profile_fields = {}
    for key in ('first_name', 'last_name', 'phone', 'avatar'):
        if key in data:
            profile_fields[f'profile.{key}'] = data[key]
    if profile_fields:
        update.update(profile_fields)

    # Role
    if 'role' in data:
        update['role'] = data['role']
        update['roles'] = [data['role']]

    if update:
        user_service.collection.update_one({'_id': ObjectId(user_id)}, {'$set': update})

    # Player fields
    player_fields_to_update = {}
    player_field_names = ['team_id', 'jersey_number', 'position', 'birth_date',
                          'height', 'weight', 'license_number', 'status']
    for field in player_field_names:
        if field in data:
            value = data[field]
            if field == 'team_id' and value:
                value = ObjectId(value)
            elif field == 'team_id' and not value:
                value = None
            elif field == 'jersey_number' and value:
                value = int(value)
            elif field == 'birth_date' and value:
                from datetime import datetime
                if isinstance(value, str):
                    value = datetime.fromisoformat(value.replace('Z', '+00:00').split('T')[0])
            elif field in ('height', 'weight') and value:
                value = int(value)
            player_fields_to_update[field] = value

    if player_fields_to_update:
        player_service = get_player_service()
        player = player_service.get_by_user(user_id)
        if player:
            player_service.collection.update_one(
                {'_id': player['_id']},
                {'$set': player_fields_to_update}
            )

    return jsonify({'success': True, 'message': 'Member updated'})


@api_bp.route('/admin/members/<user_id>/reset-password', methods=['POST'])
@role_required('admin')
def admin_reset_password(user_id):
    """Send password reset email to a member."""
    import secrets
    user_service = get_user_service()
    user = user_service.get_by_id(user_id)
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 404
    token = secrets.token_urlsafe(32)
    user_service.collection.update_one(
        {'_id': user['_id']},
        {'$set': {'reset_token': token}}
    )
    reset_link = url_for('auth.reset_password', token=token, _external=True)
    from app.services.email_service import send_reset_password_email
    success = send_reset_password_email(user['email'], reset_link)
    if success:
        return jsonify({'success': True, 'message': 'Email de réinitialisation envoyé'})
    else:
        return jsonify({'success': False, 'error': 'Erreur envoi email — vérifiez la config SMTP'}), 500


@api_bp.route('/admin/members/<user_id>', methods=['DELETE'])
@role_required('admin')
def admin_delete_member(user_id):
    """Delete a member."""
    user_service = get_user_service()
    user_service.delete(user_id)
    return jsonify({'success': True, 'message': 'Member deleted'})


@api_bp.route('/admin/check-jersey', methods=['POST'])
@role_required('admin')
def check_jersey_number():
    """Check if jersey number is available in a team."""
    data = request.get_json()
    team_id = data.get('team_id')
    jersey_number = data.get('jersey_number')
    exclude_player_id = data.get('exclude_player_id')

    if not team_id or jersey_number is None:
        return jsonify({'success': True, 'available': True})

    from bson import ObjectId
    player_service = get_player_service()
    query = {'team_id': ObjectId(team_id), 'jersey_number': int(jersey_number)}
    if exclude_player_id:
        query['_id'] = {'$ne': ObjectId(exclude_player_id)}

    existing = player_service.collection.find_one(query)
    if existing:
        return jsonify({
            'success': True,
            'available': False,
            'taken_by': existing.get('name', 'Un joueur')
        })
    return jsonify({'success': True, 'available': True})


@api_bp.route('/admin/members/<user_id>/photo', methods=['POST'])
@role_required('admin')
def admin_upload_member_photo(user_id):
    """Upload photo for a member (player)."""
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    player_service = get_player_service()
    player = player_service.get_by_user(user_id)
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in {'jpg', 'jpeg', 'png', 'webp'}:
        return jsonify({'success': False, 'error': 'Invalid file type'}), 400

    upload_dir = os.path.join(current_app.static_folder, 'uploads', 'player_docs', str(player['_id']))
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"photo.{ext}"
    file.save(os.path.join(upload_dir, filename))

    photo_url = f"/static/uploads/player_docs/{player['_id']}/{filename}"
    player_service.update(str(player['_id']), {'photo': photo_url})

    return jsonify({'success': True, 'photo': photo_url})


@api_bp.route('/admin/members/<user_id>/documents/<doc_type>', methods=['POST'])
@role_required('admin')
def admin_upload_member_document(user_id, doc_type):
    """Upload document for a member (player)."""
    if doc_type not in {'license', 'medical_cert', 'id_card', 'insurance'}:
        return jsonify({'success': False, 'error': 'Invalid document type'}), 400

    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    player_service = get_player_service()
    player = player_service.get_by_user(user_id)
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'pdf'
    if ext not in {'jpg', 'jpeg', 'png', 'pdf', 'webp'}:
        return jsonify({'success': False, 'error': 'Invalid file type'}), 400

    upload_dir = os.path.join(current_app.static_folder, 'uploads', 'player_docs', str(player['_id']))
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"{doc_type}.{ext}"
    file.save(os.path.join(upload_dir, filename))

    file_url = f"/static/uploads/player_docs/{player['_id']}/{filename}"
    player_service.update_documents(str(player['_id']), doc_type, 'valid', file_url)

    return jsonify({'success': True, 'file': file_url, 'status': 'valid'})


@api_bp.route('/admin/seed-players', methods=['POST'])
@role_required('admin')
def admin_seed_players():
    """Seed 18 demo players with French names."""
    from app.services.seed_data import seed_18_players
    data = request.get_json() or {}
    club_id = request.current_user.get('club_id')
    team_id = data.get('team_id')
    delete_existing = data.get('delete_existing', True)

    # Auto-assign to the first team if none specified
    if not team_id and club_id:
        from app.services.team_service import get_team_service
        ts = get_team_service()
        teams = ts.get_by_club(club_id)
        if teams:
            team_id = str(teams[0]['_id'])

    try:
        players = seed_18_players(club_id, team_id, delete_existing=delete_existing)
        return jsonify({'success': True, 'count': len(players), 'message': f'{len(players)} joueurs créés'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/admin/seed-all', methods=['POST'])
@role_required('admin')
def admin_seed_all():
    """Seed full demo dataset (resets all data)."""
    from app.services.seed_data import seed_all
    try:
        seed_all()
        return jsonify({'success': True, 'message': 'Données de démo injectées avec succès'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/admin/seed-coach-data', methods=['POST'])
@role_required('admin')
def admin_seed_coach_data():
    """Seed coach data (tactics, lineups, drills) WITHOUT deleting existing data."""
    from app.services.seed_data import seed_coach_data
    club_id = request.current_user.get('club_id')
    data = request.get_json() or {}
    team_id = data.get('team_id')
    try:
        seed_coach_data(club_id=club_id, team_id=team_id)
        return jsonify({'success': True, 'message': 'Données coach injectées (données existantes préservées)'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/admin/teams', methods=['POST'])
@role_required('admin')
def admin_add_team():
    """Add a new team to the club."""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'success': False, 'error': 'Team name required'}), 400
    club_id = request.current_user.get('club_id')
    team_service = get_team_service()
    team_id = team_service.create(
        club_id=club_id,
        name=data['name'],
        category=data.get('category', ''),
        colors=data.get('colors', {}),
    )
    return jsonify({'success': True, 'team_id': str(team_id)}), 201


@api_bp.route('/admin/teams/<team_id>', methods=['PUT'])
@role_required('admin')
def admin_edit_team(team_id):
    """Edit a team."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    team_service = get_team_service()
    team_service.update(team_id, data)
    return jsonify({'success': True, 'message': 'Team updated'})


@api_bp.route('/admin/teams/<team_id>', methods=['DELETE'])
@role_required('admin')
def admin_delete_team(team_id):
    """Delete a team."""
    team_service = get_team_service()
    team_service.delete(team_id)
    return jsonify({'success': True, 'message': 'Team deleted'})


@api_bp.route('/admin/teams/<team_id>/add-coach', methods=['POST'])
@role_required('admin')
def admin_add_coach_to_team(team_id):
    """Add a coach to a team."""
    data = request.get_json()
    if not data or not data.get('coach_id'):
        return jsonify({'success': False, 'error': 'coach_id required'}), 400
    team_service = get_team_service()
    team_service.add_coach(team_id, data['coach_id'])
    return jsonify({'success': True, 'message': 'Coach added to team'})


@api_bp.route('/admin/teams/<team_id>/remove-coach', methods=['POST'])
@role_required('admin')
def admin_remove_coach_from_team(team_id):
    """Remove a coach from a team."""
    data = request.get_json()
    if not data or not data.get('coach_id'):
        return jsonify({'success': False, 'error': 'coach_id required'}), 400
    team_service = get_team_service()
    team_service.remove_coach(team_id, data['coach_id'])
    return jsonify({'success': True, 'message': 'Coach removed from team'})


@api_bp.route('/admin/club', methods=['PUT'])
@role_required('admin')
def admin_update_club():
    """Update club settings."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    club_id = request.current_user.get('club_id')
    club_service = get_club_service()
    club_service.update(club_id, data)
    return jsonify({'success': True, 'message': 'Club updated'})


@api_bp.route('/admin/club/logo', methods=['POST'])
@role_required('admin')
def admin_upload_club_logo():
    """Upload the club logo and persist its URL."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'Aucun fichier fourni'}), 400

    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'Club introuvable'}), 400

    logo_url, error = _save_club_asset(club_id, file, 'logo')
    if error:
        return jsonify({'success': False, 'error': error}), 400

    club = _get_club_doc(club_id)
    personalization = _get_club_personalization(club)
    personalization['logoUrl'] = logo_url

    mongo.db.clubs.update_one(
        {'_id': ObjectId(club_id)},
        {'$set': {'logo': logo_url, 'personalization': personalization}}
    )

    return jsonify({'success': True, 'data': {'url': logo_url}})


@api_bp.route('/admin/club/cover', methods=['POST'])
@role_required('admin')
def admin_upload_club_cover():
    """Upload the club banner and persist its URL."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'Aucun fichier fourni'}), 400

    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'Club introuvable'}), 400

    cover_url, error = _save_club_asset(club_id, file, 'cover')
    if error:
        return jsonify({'success': False, 'error': error}), 400

    club = _get_club_doc(club_id)
    personalization = _get_club_personalization(club)
    personalization['coverUrl'] = cover_url

    mongo.db.clubs.update_one(
        {'_id': ObjectId(club_id)},
        {'$set': {'cover_url': cover_url, 'personalization': personalization}}
    )

    return jsonify({'success': True, 'data': {'url': cover_url}})


@api_bp.route('/admin/onboarding', methods=['GET'])
@role_required('admin')
def admin_onboarding():
    """Get onboarding status (invitations)."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})
    user_service = get_user_service()
    members = user_service.get_members_by_club(club_id)
    pending = [m for m in members if m.get('account_status') == 'pending']
    active = [m for m in members if m.get('account_status') == 'active']
    return jsonify({
        'success': True,
        'data': {
            'total': len(members),
            'pending': serialize_docs(pending),
            'active_count': len(active),
            'pending_count': len(pending),
        }
    })


@api_bp.route('/admin/analytics', methods=['GET'])
@role_required('admin')
def admin_analytics():
    """Get analytics dashboard summary."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})

    days = request.args.get('days', default=90, type=int)
    if days not in (30, 90, 365):
        days = 90

    svc = get_analytics_service()
    summary = svc.get_dashboard_summary(club_id)
    # Override growth window according to UI filter.
    summary['member_growth'] = svc.get_member_growth(club_id, days=days)
    return jsonify({'success': True, 'data': summary})


@api_bp.route('/admin/subscription', methods=['GET'])
@role_required('admin')
def admin_get_subscription():
    """Get current subscription plan."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})
    sub_service = get_subscription_service()
    sub = sub_service.get_by_club(club_id)
    return jsonify({'success': True, 'data': serialize_doc(sub) if sub else {}})


@api_bp.route('/admin/subscription', methods=['PUT'])
@role_required('admin')
def admin_update_subscription():
    """Update subscription plan."""
    data = request.get_json()
    if not data or not data.get('plan'):
        return jsonify({'success': False, 'error': 'Plan required'}), 400
    club_id = request.current_user.get('club_id')
    sub_service = get_subscription_service()
    sub_service.update_subscription(club_id, data['plan'])
    return jsonify({'success': True, 'message': 'Subscription updated'})


@api_bp.route('/admin/announcements', methods=['GET'])
@api_bp.route('/admin/subscription', methods=['DELETE'])
@role_required('admin')
def admin_cancel_subscription():
    """Cancel the club subscription."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'Club non trouvé'}), 400
    sub_service = get_subscription_service()
    sub_service.cancel_subscription(club_id)
    return jsonify({'success': True, 'message': 'Abonnement annulé'})


@api_bp.route('/admin/announcements', methods=['GET'])
@role_required('admin')
def admin_announcements():
    """Get announcement history for the admin's club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})

    svc = get_announcement_service()
    rows = svc.get_announcements(club_id)

    # Keep API payload aligned with frontend keys.
    payload = []
    for row in rows:
        target_roles = [row.get('target_id')] if row.get('target_type') == 'role' and row.get('target_id') else []
        payload.append({
            '_id': row.get('_id'),
            'title': row.get('subject', ''),
            'content': row.get('body', ''),
            'created_at': row.get('sent_at'),
            'target_roles': target_roles,
            'target_type': row.get('target_type', 'all'),
            'target_label': row.get('target_label', ''),
            'recipient_count': row.get('recipient_count', 0),
            'failed_count': row.get('failed_count', 0),
        })

    return jsonify({'success': True, 'data': serialize_docs(payload)})


@api_bp.route('/admin/announcement', methods=['POST'])
@role_required('admin')
def admin_create_announcement():
    """Create and send an announcement from frontend admin page."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    content = (data.get('content') or '').strip()
    roles = data.get('target_roles') or []

    if not title or not content:
        return jsonify({'success': False, 'error': 'title and content required'}), 400

    svc = get_announcement_service()
    sender_id = request.current_user.get('user_id')

    results = []
    if roles:
        for role in roles:
            results.append(
                svc.send_announcement(club_id, title, content, 'role', role, sender_id)
            )
    else:
        results.append(
            svc.send_announcement(club_id, title, content, 'all', '', sender_id)
        )

    total_sent = sum(r.get('recipient_count', 0) for r in results)
    total_failed = sum(r.get('failed_count', 0) for r in results)

    return jsonify({
        'success': True,
        'data': {
            'sent_count': total_sent,
            'failed_count': total_failed,
            'batches': len(results),
        }
    }), 201


@api_bp.route('/admin/email/campaigns', methods=['GET'])
@role_required('admin')
def admin_email_campaigns():
    """List email campaigns for the admin club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})

    raw = list(
        mongo.db.email_campaigns.find({'club_id': ObjectId(club_id)})
        .sort('created_at', -1)
        .limit(100)
    )
    return jsonify({'success': True, 'data': serialize_docs(raw)})


@api_bp.route('/admin/email/campaigns', methods=['POST'])
@role_required('admin')
def admin_create_email_campaign():
    """Create and optionally send an email campaign."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    subject = (data.get('subject') or '').strip()
    audience = (data.get('audience') or 'Tous les membres').strip()
    status = data.get('status') or 'sent'
    scheduled_for = data.get('scheduledFor')

    if not name or not subject:
        return jsonify({'success': False, 'error': 'name and subject required'}), 400

    target_type = 'all'
    target_id = ''
    audience_lower = audience.lower()
    if 'joueur' in audience_lower:
        target_type, target_id = 'role', 'player'
    elif 'parent' in audience_lower:
        target_type, target_id = 'role', 'parent'
    elif 'coach' in audience_lower:
        target_type, target_id = 'role', 'coach'
    elif 'supporter' in audience_lower or 'fan' in audience_lower:
        target_type, target_id = 'role', 'fan'
    elif 'admin' in audience_lower:
        target_type, target_id = 'role', 'admin'

    recipients = 0
    email_logs = []
    if status == 'sent':
        announcement_service = get_announcement_service()
        recipients_rows = announcement_service.get_announcement_recipients(club_id, target_type, target_id)
        result = announcement_service.send_announcement(
            club_id,
            subject,
            data.get('body', '') or subject,
            target_type,
            target_id,
            request.current_user.get('user_id')
        )
        recipients = result.get('recipient_count', 0)
        failed_set = set(result.get('failed_emails', []))

        now = datetime.datetime.utcnow()
        for row in recipients_rows:
            email = row.get('email')
            if not email:
                continue
            email_logs.append({
                'club_id': ObjectId(club_id),
                'to': email,
                'subject': subject,
                'campaign': name,
                'status': 'failed' if email in failed_set else 'delivered',
                'sentAt': now,
                'created_at': now,
            })

    doc = {
        'club_id': ObjectId(club_id),
        'name': name,
        'subject': subject,
        'audience': audience,
        'recipients': recipients,
        'status': status,
        'opens': 0,
        'clicks': 0,
        'bounces': 0,
        'scheduledFor': scheduled_for,
        'sentAt': datetime.datetime.utcnow() if status == 'sent' else None,
        'created_at': datetime.datetime.utcnow(),
    }
    inserted = mongo.db.email_campaigns.insert_one(doc)
    doc['_id'] = inserted.inserted_id

    if email_logs:
        mongo.db.email_logs.insert_many(email_logs)

    return jsonify({'success': True, 'data': serialize_doc(doc)}), 201


@api_bp.route('/admin/email/logs', methods=['GET'])
@role_required('admin')
def admin_email_logs():
    """List email delivery logs for the admin club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})

    query = {'club_id': ObjectId(club_id)}
    status = (request.args.get('status') or '').strip()
    q = (request.args.get('q') or '').strip()

    if status in ('delivered', 'opened', 'bounced', 'failed'):
        query['status'] = status
    if q:
        query['$or'] = [
            {'to': {'$regex': q, '$options': 'i'}},
            {'subject': {'$regex': q, '$options': 'i'}},
            {'campaign': {'$regex': q, '$options': 'i'}},
        ]

    raw = list(
        mongo.db.email_logs.find(query)
        .sort('sentAt', -1)
        .limit(300)
    )
    return jsonify({'success': True, 'data': serialize_docs(raw)})


@api_bp.route('/admin/email/templates', methods=['GET'])
@role_required('admin')
def admin_email_templates():
    """List email templates for the admin club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})

    raw = list(
        mongo.db.email_templates.find({'club_id': ObjectId(club_id)})
        .sort('updated_at', -1)
        .limit(200)
    )
    return jsonify({'success': True, 'data': serialize_docs(raw)})


@api_bp.route('/admin/email/templates', methods=['POST'])
@role_required('admin')
def admin_create_email_template():
    """Create an email template."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    subject = (data.get('subject') or '').strip()
    body = (data.get('body') or '').strip()
    category = (data.get('category') or 'announcement').strip()

    allowed_categories = {'onboarding', 'event', 'match', 'announcement', 'billing'}
    if not name or not subject or not body:
        return jsonify({'success': False, 'error': 'name, subject and body required'}), 400
    if category not in allowed_categories:
        return jsonify({'success': False, 'error': 'invalid category'}), 400

    now = datetime.datetime.utcnow()
    doc = {
        'club_id': ObjectId(club_id),
        'name': name,
        'subject': subject,
        'body': body,
        'category': category,
        'created_at': now,
        'updated_at': now,
    }
    inserted = mongo.db.email_templates.insert_one(doc)
    doc['_id'] = inserted.inserted_id
    return jsonify({'success': True, 'data': serialize_doc(doc)}), 201


@api_bp.route('/admin/email/templates/<template_id>', methods=['PUT'])
@role_required('admin')
def admin_update_email_template(template_id):
    """Update an email template."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    try:
        tpl_oid = ObjectId(template_id)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid template id'}), 400

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    subject = (data.get('subject') or '').strip()
    body = (data.get('body') or '').strip()
    category = (data.get('category') or 'announcement').strip()

    allowed_categories = {'onboarding', 'event', 'match', 'announcement', 'billing'}
    if not name or not subject or not body:
        return jsonify({'success': False, 'error': 'name, subject and body required'}), 400
    if category not in allowed_categories:
        return jsonify({'success': False, 'error': 'invalid category'}), 400

    result = mongo.db.email_templates.update_one(
        {'_id': tpl_oid, 'club_id': ObjectId(club_id)},
        {
            '$set': {
                'name': name,
                'subject': subject,
                'body': body,
                'category': category,
                'updated_at': datetime.datetime.utcnow(),
            }
        }
    )
    if result.matched_count == 0:
        return jsonify({'success': False, 'error': 'Template not found'}), 404

    updated = mongo.db.email_templates.find_one({'_id': tpl_oid})
    return jsonify({'success': True, 'data': serialize_doc(updated)})


@api_bp.route('/admin/email/templates/<template_id>', methods=['DELETE'])
@role_required('admin')
def admin_delete_email_template(template_id):
    """Delete an email template."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    try:
        tpl_oid = ObjectId(template_id)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid template id'}), 400

    result = mongo.db.email_templates.delete_one({'_id': tpl_oid, 'club_id': ObjectId(club_id)})
    if result.deleted_count == 0:
        return jsonify({'success': False, 'error': 'Template not found'}), 404

    return jsonify({'success': True, 'data': {'deleted': True}})


@api_bp.route('/admin/email/smtp', methods=['GET'])
@role_required('admin')
def admin_get_smtp_config():
    """Get SMTP configuration for the admin club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    club = mongo.db.clubs.find_one({'_id': ObjectId(club_id)}, {'smtp_config': 1}) or {}
    smtp = club.get('smtp_config') or {}
    data = {
        'host': smtp.get('host', ''),
        'port': str(smtp.get('port', '587')),
        'user': smtp.get('user', ''),
        'password': smtp.get('password', ''),
        'fromName': smtp.get('fromName', ''),
        'fromEmail': smtp.get('fromEmail', ''),
        'replyTo': smtp.get('replyTo', ''),
        'useTLS': bool(smtp.get('useTLS', True)),
    }
    return jsonify({'success': True, 'data': data})


@api_bp.route('/admin/email/smtp', methods=['PUT'])
@role_required('admin')
def admin_update_smtp_config():
    """Update SMTP configuration for the admin club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    data = request.get_json() or {}
    host = (data.get('host') or '').strip()
    port_raw = data.get('port', '587')
    user = (data.get('user') or '').strip()
    password = data.get('password') or ''
    from_name = (data.get('fromName') or '').strip()
    from_email = (data.get('fromEmail') or '').strip()
    reply_to = (data.get('replyTo') or '').strip()
    use_tls = bool(data.get('useTLS', True))

    try:
        port = int(port_raw)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'invalid port'}), 400

    if not host or not from_email:
        return jsonify({'success': False, 'error': 'host and fromEmail are required'}), 400

    smtp_config = {
        'host': host,
        'port': port,
        'user': user,
        'password': password,
        'fromName': from_name,
        'fromEmail': from_email,
        'replyTo': reply_to,
        'useTLS': use_tls,
        'updated_at': datetime.datetime.utcnow(),
    }

    mongo.db.clubs.update_one(
        {'_id': ObjectId(club_id)},
        {'$set': {'smtp_config': smtp_config, 'updated_at': datetime.datetime.utcnow()}},
    )

    smtp_config['port'] = str(smtp_config['port'])
    return jsonify({'success': True, 'data': serialize_doc(smtp_config)})


@api_bp.route('/admin/email/smtp/test', methods=['POST'])
@role_required('admin')
def admin_test_smtp_config():
    """Send a real SMTP test email using saved club SMTP settings."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    payload = request.get_json() or {}
    to_email = (payload.get('to') or '').strip()
    if not to_email or '@' not in to_email:
        return jsonify({'success': False, 'error': 'invalid recipient email'}), 400

    club = mongo.db.clubs.find_one({'_id': ObjectId(club_id)}, {'smtp_config': 1}) or {}
    smtp_saved = club.get('smtp_config') or {}
    smtp_payload = payload.get('config') if isinstance(payload.get('config'), dict) else {}
    smtp = {**smtp_saved, **smtp_payload}

    host = (smtp.get('host') or '').strip()
    user = (smtp.get('user') or '').strip()
    password = smtp.get('password') or ''
    from_name = (smtp.get('fromName') or '').strip()
    from_email = (smtp.get('fromEmail') or '').strip()
    use_tls = bool(smtp.get('useTLS', True))

    try:
        port = int(smtp.get('port', 587))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'invalid configured SMTP port'}), 400

    if not host or not from_email:
        return jsonify({'success': False, 'error': 'SMTP config incomplete: host/fromEmail required'}), 400

    subject = "FootApp - Test SMTP"
    body = "Ceci est un email de test SMTP envoyé depuis FootApp."
    msg = MIMEText(body, 'plain', 'utf-8')
    msg['Subject'] = subject
    msg['From'] = f"{from_name} <{from_email}>" if from_name else from_email
    msg['To'] = to_email

    try:
        with smtplib.SMTP(host, port, timeout=15) as server:
            server.ehlo()
            if use_tls:
                server.starttls()
                server.ehlo()
            if user:
                server.login(user, password)
            server.sendmail(from_email, [to_email], msg.as_string())
    except Exception as exc:
        current_app.logger.warning("SMTP test failed for club %s: %s", club_id, exc)

        mongo.db.email_logs.insert_one({
            'club_id': ObjectId(club_id),
            'to': to_email,
            'subject': subject,
            'campaign': 'SMTP Test',
            'status': 'failed',
            'sentAt': datetime.datetime.utcnow(),
            'created_at': datetime.datetime.utcnow(),
            'error': str(exc),
        })
        return jsonify({'success': False, 'error': 'SMTP test failed'}), 500

    mongo.db.email_logs.insert_one({
        'club_id': ObjectId(club_id),
        'to': to_email,
        'subject': subject,
        'campaign': 'SMTP Test',
        'status': 'delivered',
        'sentAt': datetime.datetime.utcnow(),
        'created_at': datetime.datetime.utcnow(),
    })

    return jsonify({'success': True, 'data': {'sent': True}})


@api_bp.route('/admin/notifications', methods=['GET'])
@role_required('admin')
def admin_notifications():
    """List admin broadcast notifications for the club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})

    raw = list(
        mongo.db.admin_notifications.find({'club_id': ObjectId(club_id)})
        .sort('created_at', -1)
        .limit(100)
    )
    return jsonify({'success': True, 'data': serialize_docs(raw)})


@api_bp.route('/admin/notifications', methods=['POST'])
@role_required('admin')
def admin_create_notification():
    """Create a broadcast notification and fan out in-app messages."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400

    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    message = (data.get('message') or '').strip()
    audience = data.get('audience') or []
    channels = data.get('channels') or ['inapp']
    priority = data.get('priority') or 'normal'
    status = data.get('status') or 'sent'
    scheduled_for = data.get('scheduledFor')

    if not title or not message:
        return jsonify({'success': False, 'error': 'title and message required'}), 400

    if not isinstance(audience, list):
        return jsonify({'success': False, 'error': 'audience must be an array'}), 400
    if not isinstance(channels, list):
        return jsonify({'success': False, 'error': 'channels must be an array'}), 400

    reach = 0
    if status == 'sent':
        user_service = get_user_service()
        notification_service = get_notification_service()
        members = user_service.get_members_by_club(club_id)
        recipients = [
            m for m in members
            if m.get('account_status') == 'active' and (not audience or m.get('role') in audience)
        ]
        for recipient in recipients:
            try:
                notification_service.create_notification(
                    str(recipient.get('_id')),
                    title,
                    message,
                    type=priority,
                )
                reach += 1
            except Exception as exc:
                current_app.logger.warning("admin notification fanout failed: %s", exc)

    doc = {
        'club_id': ObjectId(club_id),
        'title': title,
        'message': message,
        'audience': audience,
        'channels': channels,
        'priority': priority,
        'status': status,
        'scheduledFor': scheduled_for,
        'sentAt': datetime.datetime.utcnow() if status == 'sent' else None,
        'reach': reach,
        'opens': 0,
        'created_at': datetime.datetime.utcnow(),
    }
    inserted = mongo.db.admin_notifications.insert_one(doc)
    doc['_id'] = inserted.inserted_id

    return jsonify({'success': True, 'data': serialize_doc(doc)}), 201


# ============================================================
# ADMIN: ENHANCED ONBOARDING
# ============================================================

@api_bp.route('/admin/onboarding/import', methods=['POST'])
@role_required('admin')
def admin_import_csv():
    """Validate and preview CSV import."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': False, 'error': 'No club'}), 400
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'error': 'Fichier CSV requis'}), 400
    svc = get_member_onboarding_service()
    content = file.read()
    valid, errors = svc.validate_csv(content, club_id)
    return jsonify({'success': True, 'data': {'valid': valid, 'errors': errors}})


@api_bp.route('/admin/onboarding/import/confirm', methods=['POST'])
@role_required('admin')
def admin_confirm_import():
    """Confirm and execute bulk import."""
    club_id = request.current_user.get('club_id')
    data = request.get_json()
    if not data or not data.get('members'):
        return jsonify({'success': False, 'error': 'No members'}), 400
    svc = get_member_onboarding_service()
    result = svc.bulk_import_members(club_id, data['members'], data.get('custom_message'))
    return jsonify({'success': True, 'data': result})


@api_bp.route('/admin/onboarding/invitations', methods=['GET'])
@role_required('admin')
def admin_invitations():
    """Get invitation dashboard."""
    club_id = request.current_user.get('club_id')
    status_filter = request.args.get('status')
    svc = get_member_onboarding_service()
    users = svc.get_invitation_dashboard(club_id, status_filter)
    return jsonify({'success': True, 'data': serialize_docs(users)})


@api_bp.route('/admin/onboarding/resend', methods=['POST'])
@role_required('admin')
def admin_resend_invitations():
    """Resend invitations to selected members."""
    club_id = request.current_user.get('club_id')
    data = request.get_json()
    if not data or not data.get('member_ids'):
        return jsonify({'success': False, 'error': 'member_ids required'}), 400
    svc = get_member_onboarding_service()
    count = svc.resend_invitations(club_id, data['member_ids'])
    return jsonify({'success': True, 'resent_count': count})


# ============================================================
# ADMIN: ENHANCED ANALYTICS
# ============================================================

@api_bp.route('/admin/analytics/teams', methods=['GET'])
@role_required('admin')
def admin_analytics_teams():
    """Get team performance analytics."""
    club_id = request.current_user.get('club_id')
    svc = get_analytics_service()
    return jsonify({'success': True, 'data': svc.get_team_performance(club_id)})


@api_bp.route('/admin/analytics/retention', methods=['GET'])
@role_required('admin')
def admin_analytics_retention():
    """Get member retention metrics."""
    club_id = request.current_user.get('club_id')
    days = request.args.get('days', default=90, type=int)
    if days not in (30, 90, 365):
        days = 90
    svc = get_analytics_service()
    return jsonify({'success': True, 'data': svc.get_member_retention(club_id, days=days)})


@api_bp.route('/admin/analytics/engagement', methods=['GET'])
@role_required('admin')
def admin_analytics_engagement():
    """Get feature usage metrics."""
    club_id = request.current_user.get('club_id')
    svc = get_analytics_service()
    return jsonify({'success': True, 'data': svc.get_feature_usage(club_id)})


@api_bp.route('/admin/analytics/financial', methods=['GET'])
@role_required('admin')
def admin_analytics_financial():
    """Get financial analytics."""
    club_id = request.current_user.get('club_id')
    svc = get_analytics_service()
    return jsonify({'success': True, 'data': svc.get_financial_metrics(club_id)})


@api_bp.route('/admin/analytics/export/pdf', methods=['GET'])
@role_required('admin')
def admin_analytics_export_pdf():
    """Export analytics summary as PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    club_id = request.current_user.get('club_id')
    svc = get_analytics_service()
    summary = svc.get_dashboard_summary(club_id)
    financial = svc.get_financial_metrics(club_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('Analytics — Rapport du club', styles['Title']))
    story.append(Spacer(1, 12))

    # KPI table
    kpi_data = [['Indicateur', 'Valeur']]
    kpi_data.append(['Total membres', str(summary.get('total_members', 0))])
    kpi_data.append(['Membres actifs', str(summary.get('active_members', 0))])
    kpi_data.append(['Équipes', str(summary.get('total_teams', 0))])
    kpi_data.append(['Revenus (€)', str(financial.get('total_revenue', 0))])
    kpi_data.append(['Dépenses (€)', str(financial.get('total_expenses', 0))])
    kpi_data.append(['Bénéfice net (€)', str(financial.get('net_profit', 0))])

    t = Table(kpi_data, colWidths=[250, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22c55e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f9fafb'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    doc.build(story)

    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=analytics.pdf'
    return response


@api_bp.route('/admin/analytics/export/excel', methods=['GET'])
@role_required('admin')
def admin_analytics_export_excel():
    """Export analytics summary as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    club_id = request.current_user.get('club_id')
    svc = get_analytics_service()
    summary = svc.get_dashboard_summary(club_id)
    financial = svc.get_financial_metrics(club_id)
    growth = svc.get_member_growth(club_id, days=90)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Analytics'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='22C55E')

    headers = ['Indicateur', 'Valeur']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    rows = [
        ('Total membres', summary.get('total_members', 0)),
        ('Membres actifs', summary.get('active_members', 0)),
        ('Équipes', summary.get('total_teams', 0)),
        ('Revenus (€)', financial.get('total_revenue', 0)),
        ('Dépenses (€)', financial.get('total_expenses', 0)),
        ('Bénéfice net (€)', financial.get('net_profit', 0)),
    ]
    for r, (label, value) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # Growth sheet
    if growth:
        ws2 = wb.create_sheet('Croissance')
        ws2.append(['Période', 'Nouveaux membres'])
        for g in growth:
            ws2.append([g.get('period', ''), g.get('new_members', 0)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=analytics.xlsx'
    return response


@api_bp.route('/admin/personalization', methods=['GET'])
@role_required('admin')
def admin_get_personalization():
    """Get club personalization settings."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})
    club = _get_club_doc(club_id)
    if not club:
        return jsonify({'success': True, 'data': {}})
    data = _get_club_personalization(club)
    return jsonify({'success': True, 'data': data})


@api_bp.route('/admin/personalization', methods=['PUT'])
@role_required('admin')
def admin_update_personalization():
    """Save club personalization settings."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    club_id = request.current_user.get('club_id')
    club = _get_club_doc(club_id)
    personalization = _get_club_personalization(club)
    personalization.update(data)

    update = {'personalization': personalization}
    # Sync top-level name field if provided
    if 'clubName' in personalization:
        update['name'] = personalization['clubName']
    if 'logoUrl' in personalization:
        update['logo'] = personalization['logoUrl']
    if 'coverUrl' in personalization:
        update['cover_url'] = personalization['coverUrl']
    if 'primaryColor' in personalization or 'accentColor' in personalization:
        update['colors'] = {
            'primary': personalization.get('primaryColor', '#22c55e'),
            'secondary': personalization.get('accentColor', '#16a34a'),
        }
    mongo.db.clubs.update_one({'_id': ObjectId(club_id)}, {'$set': update})
    return jsonify({'success': True, 'message': 'Personalization saved'})

@api_bp.route('/admin/billing/dashboard', methods=['GET'])
@role_required('admin')
def admin_billing_dashboard():
    """Get billing dashboard."""
    club_id = request.current_user.get('club_id')
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_billing_dashboard(club_id)})


@api_bp.route('/admin/billing/invoices', methods=['GET'])
@role_required('admin')
def admin_billing_invoices():
    """Get invoices."""
    club_id = request.current_user.get('club_id')
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_invoices(club_id)})


@api_bp.route('/admin/billing/invoices/<invoice_id>/pdf', methods=['GET'])
@role_required('admin')
def admin_billing_invoice_pdf(invoice_id):
    """Download a PDF version of an invoice."""
    svc = get_billing_service()
    pdf_bytes = svc.generate_invoice_pdf(invoice_id)
    if not pdf_bytes:
        return jsonify({'success': False, 'error': 'Invoice not found or PDF unavailable'}), 404

    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=invoice-{invoice_id}.pdf'
    return response


# ============================================================
# PARENT ENDPOINTS
# ============================================================

@api_bp.route('/parent/dashboard', methods=['GET'])
@role_required('parent')
def parent_dashboard():
    """Get parent dashboard with linked children."""
    user_id = request.current_user['user_id']
    parent_link_service = get_parent_link_service()
    player_service = get_player_service()

    links = parent_link_service.get_linked_players(user_id)
    children = []
    for link in links:
        player = player_service.get_by_id(link.get('player_id'))
        if player:
            user_service = get_user_service()
            user = user_service.get_by_id(str(player.get('user_id', '')))
            child_data = serialize_doc(player)
            if user:
                child_data['user_profile'] = user.get('profile', {})
            children.append(child_data)

    return jsonify({'success': True, 'data': {'children': children}})


@api_bp.route('/parent/link', methods=['POST'])
@role_required('parent')
def parent_link_child():
    """Link child to parent via 6-char code."""
    data = request.get_json()
    code = data.get('code', '').strip() if data else ''
    if not code or len(code) != 6:
        return jsonify({'success': False, 'error': 'Valid 6-character code required'}), 400

    parent_link_service = get_parent_link_service()
    result = parent_link_service.link_parent_to_player(request.current_user['user_id'], code)
    if result:
        return jsonify({'success': True, 'message': 'Child linked successfully'})
    return jsonify({'success': False, 'error': 'Invalid or expired code'}), 400


@api_bp.route('/parent/children/<player_id>/calendar', methods=['GET'])
@role_required('parent')
def parent_child_calendar(player_id):
    """Get calendar for a linked child's team."""
    player_service = get_player_service()
    player = player_service.get_by_id(player_id)
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404

    team_id = player.get('team_id')
    club_id = player.get('club_id')
    event_service = get_event_service()
    match_service = get_match_service()
    events = event_service.get_upcoming(str(club_id), team_id=str(team_id) if team_id else None, limit=20)
    matches = match_service.get_upcoming(str(club_id), team_id=str(team_id) if team_id else None, limit=10)
    return jsonify({
        'success': True,
        'data': {
            'events': serialize_docs(events),
            'matches': serialize_docs(matches),
        }
    })


@api_bp.route('/parent/children/<player_id>/roster', methods=['GET'])
@role_required('parent')
def parent_child_roster(player_id):
    """Get roster for a child's team (restricted view)."""
    player_service = get_player_service()
    player = player_service.get_by_id(player_id)
    if not player:
        return jsonify({'success': False, 'error': 'Player not found'}), 404

    team_id = player.get('team_id')
    if not team_id:
        return jsonify({'success': True, 'data': []})

    team_service = get_team_service()
    players = team_service.get_players(str(team_id))
    safe_data = []
    for p in players:
        safe_data.append({
            '_id': str(p.get('_id', '')),
            'jersey_number': p.get('jersey_number'),
            'position': p.get('position'),
            'first_name': p.get('first_name', ''),
            'last_name': p.get('last_name', ''),
        })
    return jsonify({'success': True, 'data': safe_data})


@api_bp.route('/parent/generate-code/<player_id>', methods=['POST'])
@role_required('coach', 'admin')
def generate_parent_code(player_id):
    """Generate a parent linking code for a player."""
    parent_link_service = get_parent_link_service()
    code = parent_link_service.generate_link_code(player_id)
    return jsonify({'success': True, 'code': code})


# ============================================================
# PARENT: ENHANCED ENDPOINTS
# ============================================================

@api_bp.route('/parent/children/<player_id>/progress', methods=['GET'])
@role_required('parent')
def parent_child_progress(player_id):
    """Get child progress monitoring data."""
    svc = get_parent_monitoring_service()
    data = svc.get_child_progress(player_id)
    if not data:
        return jsonify({'success': False, 'error': 'Joueur non trouvé'}), 404
    return jsonify({'success': True, 'data': data})


@api_bp.route('/parent/children/<player_id>/feedback', methods=['GET'])
@role_required('parent')
def parent_child_feedback(player_id):
    """Get coach feedback for child."""
    svc = get_parent_monitoring_service()
    data = svc.get_coach_feedback(player_id)
    return jsonify({'success': True, 'data': data})


@api_bp.route('/parent/children/<player_id>/achievements', methods=['GET'])
@role_required('parent')
def parent_child_achievements(player_id):
    """Get child achievements."""
    svc = get_parent_monitoring_service()
    data = svc.get_achievements(player_id)
    return jsonify({'success': True, 'data': data})


@api_bp.route('/parent/messages/coach/<coach_id>', methods=['GET'])
@role_required('parent')
def parent_coach_messages(coach_id):
    """Get parent-coach message thread."""
    user_id = request.current_user['user_id']
    svc = MessagingService(mongo.db)
    messages = svc.get_direct_messages(user_id, coach_id)
    return jsonify({'success': True, 'data': serialize_docs(messages)})


@api_bp.route('/parent/absence-report', methods=['POST'])
@role_required('parent')
def parent_absence_report():
    """Report child absence."""
    data = request.get_json()
    if not data or not data.get('player_id') or not data.get('event_id'):
        return jsonify({'success': False, 'error': 'player_id et event_id requis'}), 400
    event_service = get_event_service()
    event_service.update_attendance(data['event_id'], data['player_id'], 'absent', data.get('reason', ''))
    return jsonify({'success': True, 'message': 'Absence signalée'})


@api_bp.route('/parent/payments', methods=['GET'])
@role_required('parent')
def parent_payments():
    """Get parent payment records."""
    user_id = request.current_user['user_id']
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_parent_payments(user_id)})


@api_bp.route('/parent/payments/categories', methods=['GET'])
@role_required('parent')
def parent_payment_categories():
    """Get payment categories."""
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_payment_categories()})


# ============================================================
# ISY ENDPOINTS
@api_bp.route('/parent/coaches', methods=['GET'])
@role_required('parent')
def parent_coaches():
    """Get coaches accessible via linked children."""
    user_id = request.current_user['user_id']
    from bson import ObjectId as ObjId
    # Find linked players
    links = list(mongo.db.parent_links.find({'parent_user_id': user_id}))
    player_ids = [l.get('player_id') for l in links if l.get('player_id')]
    coaches = []
    seen = set()
    for pid in player_ids:
        player = mongo.db.players.find_one({'_id': ObjId(str(pid)) if isinstance(pid, str) else pid})
        if not player:
            continue
        team_id = player.get('team_id')
        if not team_id:
            continue
        team = mongo.db.teams.find_one({'_id': team_id})
        if not team:
            continue
        coach_user_id = team.get('coach_user_id') or team.get('coach_id')
        if not coach_user_id or str(coach_user_id) in seen:
            continue
        seen.add(str(coach_user_id))
        cu = mongo.db.users.find_one({'_id': ObjId(str(coach_user_id)) if isinstance(coach_user_id, str) else coach_user_id})
        if cu:
            coaches.append({
                'id': str(cu['_id']),
                'name': cu.get('full_name') or (cu.get('first_name', '') + ' ' + cu.get('last_name', '')).strip(),
                'player_name': player.get('name', ''),
                'team_name': team.get('name', ''),
            })
    return jsonify({'success': True, 'data': coaches})


@api_bp.route('/parent/messages/coach/<coach_id>', methods=['POST'])
@role_required('parent')
def parent_send_coach_message(coach_id):
    """Send a message from parent to coach."""
    data = request.get_json()
    if not data or not data.get('content'):
        return jsonify({'success': False, 'error': 'content requis'}), 400
    user_id = request.current_user['user_id']
    svc = MessagingService(mongo.db)
    svc.send_direct_message(user_id, coach_id, data['content'])
    return jsonify({'success': True, 'message': 'Message envoyé'})


@api_bp.route('/parent/payments/export/csv', methods=['GET'])
@role_required('parent')
def parent_payments_export_csv():
    """Export parent payments to Excel."""
    import io
    try:
        import openpyxl
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    user_id = request.current_user['user_id']
    svc = get_billing_service()
    payments = svc.get_parent_payments(user_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Paiements'
    ws.append(['Description', 'Catégorie', 'Montant (€)', 'Statut', 'Échéance', 'Payé le'])
    for p in (payments or []):
        ws.append([
            p.get('description', ''),
            p.get('category', ''),
            p.get('amount', 0),
            p.get('status', ''),
            str(p.get('due_date', '')),
            str(p.get('paid_date', '') or ''),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='paiements.xlsx',
    )


# ============================================================
# ISY ENDPOINTS
# ============================================================

@api_bp.route('/isy/dashboard', methods=['GET'])
@role_required('coach', 'admin')
def isy_dashboard():
    """Get ISY hub dashboard."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': {}})
    sponsors = list(mongo.db.sponsors.find({'club_id': club_id}))
    payments = list(mongo.db.payments.find({'club_id': club_id}).sort('date', -1).limit(20))
    total_revenue = sum(p.get('amount', 0) for p in payments if p.get('status') == 'confirmed')
    return jsonify({
        'success': True,
        'data': {
            'sponsors': serialize_docs(sponsors),
            'recent_payments': serialize_docs(payments),
            'total_sponsors': len(sponsors),
            'total_revenue': total_revenue,
        }
    })


@api_bp.route('/isy/sponsors', methods=['GET'])
@role_required('coach', 'admin')
def isy_get_sponsors():
    """Get sponsors list."""
    club_id = request.current_user.get('club_id')
    sponsors = list(mongo.db.sponsors.find({'club_id': club_id}))
    return jsonify({'success': True, 'data': serialize_docs(sponsors)})


@api_bp.route('/isy/sponsors', methods=['POST'])
@role_required('coach', 'admin')
def isy_add_sponsor():
    """Add a new sponsor."""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'success': False, 'error': 'Sponsor name required'}), 400
    club_id = request.current_user.get('club_id')
    sponsor_id = mongo.db.sponsors.insert_one({
        'club_id': club_id,
        'name': data['name'],
        'contact': data.get('contact', ''),
        'amount': data.get('amount', 0),
        'type': data.get('type', 'gold'),
        'created_at': datetime.datetime.utcnow(),
    }).inserted_id
    return jsonify({'success': True, 'sponsor_id': str(sponsor_id)}), 201


@api_bp.route('/isy/sponsors/<sponsor_id>', methods=['DELETE'])
@role_required('coach', 'admin')
def isy_delete_sponsor(sponsor_id):
    """Delete a sponsor."""
    mongo.db.sponsors.delete_one({'_id': ObjectId(sponsor_id)})
    return jsonify({'success': True, 'message': 'Sponsor deleted'})


@api_bp.route('/isy/payments', methods=['GET'])
@role_required('coach', 'admin')
def isy_get_payments():
    """Get payments list."""
    club_id = request.current_user.get('club_id')
    payments = list(mongo.db.payments.find({'club_id': club_id}).sort('date', -1))
    return jsonify({'success': True, 'data': serialize_docs(payments)})


@api_bp.route('/isy/payments', methods=['POST'])
@role_required('coach', 'admin')
def isy_add_payment():
    """Record a payment."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400
    club_id = request.current_user.get('club_id')
    payment_id = mongo.db.payments.insert_one({
        'club_id': club_id,
        'player_id': data.get('player_id'),
        'player_name': data.get('player_name', ''),
        'amount': data.get('amount', 0),
        'type': data.get('type', 'cotisation'),
        'status': data.get('status', 'pending'),
        'date': datetime.datetime.utcnow(),
        'description': data.get('description', ''),
    }).inserted_id
    return jsonify({'success': True, 'payment_id': str(payment_id)}), 201


@api_bp.route('/isy/payments/<payment_id>/confirm', methods=['POST'])
@role_required('coach', 'admin')
def isy_confirm_payment(payment_id):
    """Confirm a payment."""
    mongo.db.payments.update_one(
        {'_id': ObjectId(payment_id)},
        {'$set': {'status': 'confirmed'}}
    )
    return jsonify({'success': True, 'message': 'Payment confirmed'})


# ============================================================
# SHOP ENDPOINTS
# ============================================================

@api_bp.route('/shop/products', methods=['GET'])
def shop_get_products():
    """Get shop products."""
    club_id = request.args.get('club_id')
    query = {'club_id': club_id} if club_id else {}
    products = list(mongo.db.products.find(query))
    return jsonify({'success': True, 'data': serialize_docs(products)})


@api_bp.route('/shop/products/<product_id>', methods=['GET'])
def shop_get_product(product_id):
    """Get single product."""
    product = mongo.db.products.find_one({'_id': ObjectId(product_id)})
    if not product:
        return jsonify({'success': False, 'error': 'Product not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(product)})


@api_bp.route('/shop/orders', methods=['GET'])
@token_required
def shop_get_orders():
    """Get user's orders."""
    user_id = request.current_user['user_id']
    orders = list(mongo.db.orders.find({'user_id': user_id}).sort('created_at', -1))
    return jsonify({'success': True, 'data': serialize_docs(orders)})


@api_bp.route('/shop/orders', methods=['POST'])
@token_required
def shop_create_order():
    """Create a new order."""
    data = request.get_json()
    if not data or not data.get('items'):
        return jsonify({'success': False, 'error': 'Items required'}), 400

    user_id = request.current_user['user_id']
    club_id = request.current_user.get('club_id')
    items = data['items']
    total = sum(item.get('price', 0) * item.get('quantity', 1) for item in items)

    order_id = mongo.db.orders.insert_one({
        'user_id': user_id,
        'club_id': club_id,
        'items': items,
        'total': total,
        'status': 'pending',
        'created_at': datetime.datetime.utcnow(),
    }).inserted_id
    return jsonify({'success': True, 'order_id': str(order_id)}), 201


# ============================================================
# SHOP CATEGORIES ENDPOINT
# ============================================================

@api_bp.route('/shop/categories', methods=['GET'])
def shop_get_categories():
    """Get product categories."""
    products = list(mongo.db.products.find({}, {'category': 1}))
    categories = list(set(p.get('category', 'Autre') for p in products if p.get('category')))
    return jsonify({'success': True, 'data': sorted(categories)})


# ============================================================
# COMPETITIONS ENDPOINTS
# ============================================================

@api_bp.route('/competitions', methods=['GET'])
def get_competitions():
    """Get competitions."""
    club_id = request.args.get('club_id')
    query = {'club_id': club_id} if club_id else {}
    competitions = list(mongo.db.competitions.find(query).sort('start_date', -1))
    return jsonify({'success': True, 'data': serialize_docs(competitions)})


@api_bp.route('/competitions/<competition_id>', methods=['GET'])
def get_competition(competition_id):
    """Get single competition."""
    comp = mongo.db.competitions.find_one({'_id': ObjectId(competition_id)})
    if not comp:
        return jsonify({'success': False, 'error': 'Competition not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(comp)})


# ============================================================
# COACH SCOUTING ENDPOINT
# ============================================================

@api_bp.route('/coach/scouting', methods=['GET'])
@role_required('coach')
def coach_scouting():
    """Get scouting prospects."""
    user_id = request.current_user['user_id']
    club_id = request.current_user.get('club_id')
    prospects = list(mongo.db.scouting.find({'club_id': club_id}).sort('created_at', -1))
    return jsonify({'success': True, 'data': serialize_docs(prospects)})


@api_bp.route('/coach/scouting', methods=['POST'])
@role_required('coach')
def coach_add_prospect():
    """Add a scouting prospect."""
    data = request.get_json()
    if not data or not data.get('first_name'):
        return jsonify({'success': False, 'error': 'First name required'}), 400
    club_id = request.current_user.get('club_id')
    prospect_id = mongo.db.scouting.insert_one({
        'club_id': club_id,
        'first_name': data['first_name'],
        'last_name': data.get('last_name', ''),
        'position': data.get('position', ''),
        'club': data.get('club', ''),
        'notes': data.get('notes', ''),
        'rating': data.get('rating'),
        'created_at': datetime.datetime.utcnow(),
    }).inserted_id
    return jsonify({'success': True, 'prospect_id': str(prospect_id)}), 201


# ============================================================
# COACH TRAINING PLANS ENDPOINTS
# ============================================================

@api_bp.route('/coach/training-plans', methods=['GET'])
@role_required('coach')
def coach_training_plans():
    """Get training plans for team."""
    from app.services import get_training_service
    team_id = request.args.get('team_id') or request.current_user.get('team_id')
    if not team_id:
        club_id = request.current_user.get('club_id')
        team = mongo.db.teams.find_one({'club_id': ObjectId(club_id)}) if club_id else None
        team_id = str(team['_id']) if team else None
    if not team_id:
        return jsonify({'success': True, 'data': []})
    svc = get_training_service()
    plans = svc.get_plans(team_id, status=request.args.get('status'))
    return jsonify({'success': True, 'data': serialize_docs(plans)})


@api_bp.route('/coach/training-plans', methods=['POST'])
@role_required('coach')
def coach_create_training_plan():
    """Create a training plan."""
    from app.services import get_training_service
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'success': False, 'error': 'Name required'}), 400
    club_id = request.current_user.get('club_id')
    team_id = data.get('team_id') or request.args.get('team_id')
    if not team_id:
        team = mongo.db.teams.find_one({'club_id': ObjectId(club_id)}) if club_id else None
        team_id = str(team['_id']) if team else None
    if not team_id:
        return jsonify({'success': False, 'error': 'Team required'}), 400
    svc = get_training_service()
    plan_id = svc.create_plan(club_id, team_id, request.current_user['user_id'], data)
    return jsonify({'success': True, 'plan_id': plan_id}), 201


@api_bp.route('/coach/training-plans/<plan_id>', methods=['GET'])
@role_required('coach')
def coach_training_plan_detail(plan_id):
    """Get a single training plan with its sessions."""
    from app.services import get_training_service
    svc = get_training_service()
    plan = svc.get_plan(plan_id)
    if not plan:
        return jsonify({'success': False, 'error': 'Plan not found'}), 404
    sessions = svc.get_sessions(plan_id=plan_id)
    result = serialize_doc(plan)
    result['sessions'] = serialize_docs(sessions)
    return jsonify({'success': True, 'data': result})


@api_bp.route('/coach/training-plans/<plan_id>', methods=['PUT'])
@role_required('coach')
def coach_update_training_plan(plan_id):
    """Update a training plan."""
    from app.services import get_training_service
    data = request.get_json()
    svc = get_training_service()
    svc.update_plan(plan_id, data)
    return jsonify({'success': True})


@api_bp.route('/coach/training-plans/<plan_id>', methods=['DELETE'])
@role_required('coach')
def coach_delete_training_plan(plan_id):
    """Delete a training plan and its sessions."""
    from app.services import get_training_service
    svc = get_training_service()
    svc.delete_plan(plan_id)
    return jsonify({'success': True})


@api_bp.route('/coach/training-plans/<plan_id>/sessions', methods=['POST'])
@role_required('coach')
def coach_create_session(plan_id):
    """Add a session to a training plan."""
    from app.services import get_training_service
    data = request.get_json()
    svc = get_training_service()
    session_id = svc.create_session(plan_id, request.current_user['user_id'], data or {})
    if not session_id:
        return jsonify({'success': False, 'error': 'Plan not found'}), 404
    return jsonify({'success': True, 'session_id': session_id}), 201


@api_bp.route('/coach/training-sessions/<session_id>', methods=['GET'])
@role_required('coach')
def coach_session_detail(session_id):
    """Get a training session."""
    from app.services import get_training_service
    svc = get_training_service()
    session = svc.get_session(session_id)
    if not session:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(session)})


@api_bp.route('/coach/training-sessions/<session_id>', methods=['PUT'])
@role_required('coach')
def coach_update_session(session_id):
    """Update a training session."""
    from app.services import get_training_service
    data = request.get_json()
    svc = get_training_service()
    svc.update_session(session_id, data or {})
    return jsonify({'success': True})


@api_bp.route('/coach/training-sessions/<session_id>/attendance', methods=['POST'])
@role_required('coach')
def coach_session_attendance(session_id):
    """Mark attendance for a training session."""
    from app.services import get_training_service
    data = request.get_json()
    svc = get_training_service()
    if isinstance(data, list):
        svc.bulk_attendance(session_id, data)
    elif data and data.get('player_id'):
        svc.mark_attendance(session_id, data['player_id'], data.get('status', 'present'),
                            data.get('reason'), data.get('rating'))
    return jsonify({'success': True})


@api_bp.route('/coach/drills', methods=['GET'])
@role_required('coach')
def coach_drills():
    """Get drill library."""
    from app.services import get_training_service
    club_id = request.current_user.get('club_id')
    svc = get_training_service()
    drills = svc.get_drills(
        club_id=club_id,
        category=request.args.get('category'),
        difficulty=request.args.get('difficulty'),
    )
    return jsonify({'success': True, 'data': serialize_docs(drills)})


@api_bp.route('/coach/drills', methods=['POST'])
@role_required('coach')
def coach_create_drill():
    """Create a custom drill."""
    from app.services import get_training_service
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'success': False, 'error': 'Name required'}), 400
    club_id = request.current_user.get('club_id')
    svc = get_training_service()
    drill_id = svc.create_drill(club_id, request.current_user['user_id'], data)
    return jsonify({'success': True, 'drill_id': drill_id}), 201


@api_bp.route('/coach/drills/<drill_id>', methods=['GET'])
@role_required('coach')
def coach_drill_detail(drill_id):
    """Get drill details."""
    from app.services import get_training_service
    svc = get_training_service()
    drill = svc.get_drill(drill_id)
    if not drill:
        return jsonify({'success': False, 'error': 'Drill not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(drill)})


@api_bp.route('/coach/training-load/<player_id>', methods=['GET'])
@role_required('coach')
def coach_training_load(player_id):
    """Get training load for a player."""
    from app.services import get_training_service
    weeks = int(request.args.get('weeks', 1))
    svc = get_training_service()
    load = svc.get_training_load(player_id, weeks)
    return jsonify({'success': True, 'data': load})


# ============================================================
# COACH INJURY ENDPOINTS
# ============================================================

@api_bp.route('/coach/injuries', methods=['GET'])
@role_required('coach')
def coach_injuries():
    """Get injuries for team."""
    from app.services import get_injury_service
    team_id = request.args.get('team_id')
    if not team_id:
        club_id = request.current_user.get('club_id')
        team = mongo.db.teams.find_one({'club_id': ObjectId(club_id)}) if club_id else None
        team_id = str(team['_id']) if team else None
    if not team_id:
        return jsonify({'success': True, 'data': []})
    svc = get_injury_service()
    injuries = svc.get_injuries(team_id, status=request.args.get('status'))
    # Enrich with player names
    for inj in injuries:
        player = mongo.db.players.find_one({'_id': inj.get('player_id')})
        inj['player_name'] = player.get('name', '') if player else ''
    return jsonify({'success': True, 'data': serialize_docs(injuries)})


@api_bp.route('/coach/injuries', methods=['POST'])
@role_required('coach')
def coach_log_injury():
    """Log a new injury."""
    from app.services import get_injury_service
    data = request.get_json()
    if not data or not data.get('player_id'):
        return jsonify({'success': False, 'error': 'Player ID required'}), 400
    team_id = data.get('team_id')
    if not team_id:
        player = mongo.db.players.find_one({'_id': ObjectId(data['player_id'])})
        team_id = str(player['team_id']) if player and player.get('team_id') else None
    if not team_id:
        return jsonify({'success': False, 'error': 'Team not found'}), 400
    svc = get_injury_service()
    injury_id = svc.log_injury(data['player_id'], request.current_user['user_id'], team_id, data)
    return jsonify({'success': True, 'injury_id': injury_id}), 201


@api_bp.route('/coach/injuries/<injury_id>', methods=['GET'])
@role_required('coach')
def coach_injury_detail(injury_id):
    """Get injury details."""
    from app.services import get_injury_service
    svc = get_injury_service()
    injury = svc.get_injury(injury_id)
    if not injury:
        return jsonify({'success': False, 'error': 'Injury not found'}), 404
    player = mongo.db.players.find_one({'_id': injury.get('player_id')})
    result = serialize_doc(injury)
    result['player_name'] = player.get('name', '') if player else ''
    return jsonify({'success': True, 'data': result})


@api_bp.route('/coach/injuries/<injury_id>', methods=['PUT'])
@role_required('coach')
def coach_update_injury(injury_id):
    """Update injury recovery notes."""
    from app.services import get_injury_service
    data = request.get_json()
    svc = get_injury_service()
    svc.update_recovery(injury_id, request.current_user['user_id'], data or {})
    return jsonify({'success': True})


@api_bp.route('/coach/injuries/<injury_id>/clear', methods=['POST'])
@role_required('coach')
def coach_clear_injury(injury_id):
    """Clear a player for play."""
    from app.services import get_injury_service
    data = request.get_json() or {}
    svc = get_injury_service()
    svc.clear_for_play(injury_id, data.get('cleared_by', 'Coach'), data.get('date'))
    return jsonify({'success': True})


@api_bp.route('/coach/injuries/stats', methods=['GET'])
@role_required('coach')
def coach_injury_stats():
    """Get injury statistics for team."""
    from app.services import get_injury_service
    team_id = request.args.get('team_id')
    if not team_id:
        club_id = request.current_user.get('club_id')
        team = mongo.db.teams.find_one({'club_id': ObjectId(club_id)}) if club_id else None
        team_id = str(team['_id']) if team else None
    if not team_id:
        return jsonify({'success': True, 'data': {}})
    svc = get_injury_service()
    stats = svc.get_injury_stats(team_id)
    return jsonify({'success': True, 'data': stats})


@api_bp.route('/coach/injuries/player/<player_id>', methods=['GET'])
@role_required('coach')
def coach_player_injuries(player_id):
    """Get injury history for a player."""
    from app.services import get_injury_service
    svc = get_injury_service()
    injuries = svc.get_player_injuries(player_id)
    return jsonify({'success': True, 'data': serialize_docs(injuries)})


# ============================================================
# COACH PLAYER ANALYTICS ENDPOINTS
# ============================================================

@api_bp.route('/coach/analytics/players', methods=['GET'])
@role_required('coach')
def coach_analytics_players():
    """Get team rankings / player list for analytics."""
    from app.services import get_player_analytics_service
    team_id = request.args.get('team_id')
    if not team_id:
        club_id = request.current_user.get('club_id')
        team = mongo.db.teams.find_one({'club_id': ObjectId(club_id)}) if club_id else None
        team_id = str(team['_id']) if team else None
    if not team_id:
        return jsonify({'success': True, 'data': []})
    svc = get_player_analytics_service()
    rankings = svc.get_team_rankings(team_id)
    return jsonify({'success': True, 'data': rankings})


@api_bp.route('/coach/analytics/player/<player_id>', methods=['GET'])
@role_required('coach')
def coach_analytics_player_detail(player_id):
    """Get comprehensive player dashboard."""
    from app.services import get_player_analytics_service
    svc = get_player_analytics_service()
    dashboard = svc.get_player_dashboard(player_id)
    if not dashboard:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    return jsonify({'success': True, 'data': dashboard})


@api_bp.route('/coach/analytics/compare', methods=['POST'])
@role_required('coach')
def coach_analytics_compare():
    """Compare 2-5 players side-by-side."""
    from app.services import get_player_analytics_service
    data = request.get_json()
    if not data or not data.get('player_ids'):
        return jsonify({'success': False, 'error': 'player_ids required'}), 400
    svc = get_player_analytics_service()
    comparison = svc.compare_players(data['player_ids'])
    return jsonify({'success': True, 'data': comparison})


@api_bp.route('/coach/analytics/player/<player_id>/trends', methods=['GET'])
@role_required('coach')
def coach_analytics_trends(player_id):
    """Get trend analysis for a player."""
    from app.services import get_player_analytics_service
    svc = get_player_analytics_service()
    trends = svc.get_trend_analysis(player_id)
    if not trends:
        return jsonify({'success': False, 'error': 'Player not found'}), 404
    return jsonify({'success': True, 'data': trends})


# ============================================================
# SUPERADMIN ENDPOINTS
# ============================================================

@api_bp.route('/superadmin/dashboard', methods=['GET'])
@role_required('superadmin')
def superadmin_dashboard():
    """Get platform-wide dashboard."""
    clubs = list(mongo.db.clubs.find())
    users = mongo.db.users.count_documents({})
    players = mongo.db.players.count_documents({})
    return jsonify({
        'success': True,
        'data': {
            'total_clubs': len(clubs),
            'total_users': users,
            'total_players': players,
            'clubs': serialize_docs(clubs),
        }
    })


@api_bp.route('/superadmin/projects', methods=['GET'])
@role_required('superadmin')
def superadmin_projects():
    """Get all projects."""
    project_service = get_project_service()
    projects = project_service.get_all()
    return jsonify({'success': True, 'data': serialize_docs(projects)})


@api_bp.route('/superadmin/projects', methods=['POST'])
@role_required('superadmin')
def superadmin_create_project():
    """Create a new project."""
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'success': False, 'error': 'Name required'}), 400
    project_service = get_project_service()
    project_id = project_service.create(data)
    return jsonify({'success': True, 'project_id': str(project_id)}), 201


@api_bp.route('/superadmin/projects/<project_id>', methods=['GET'])
@role_required('superadmin')
def superadmin_project_detail(project_id):
    """Get project details with tickets."""
    project_service = get_project_service()
    project = project_service.get_by_id(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    return jsonify({'success': True, 'data': serialize_doc(project)})


@api_bp.route('/superadmin/projects/<project_id>/tickets', methods=['POST'])
@role_required('superadmin')
def superadmin_create_ticket(project_id):
    """Create a ticket for a project."""
    data = request.get_json()
    if not data or not data.get('title'):
        return jsonify({'success': False, 'error': 'Title required'}), 400
    project_service = get_project_service()
    ticket_id = project_service.create_ticket(project_id, data)
    return jsonify({'success': True, 'ticket_id': str(ticket_id)}), 201


@api_bp.route('/superadmin/clubs', methods=['GET'])
@role_required('superadmin')
def superadmin_clubs():
    """Get all clubs for management."""
    clubs = list(mongo.db.clubs.find())
    result = []
    for club in clubs:
        user_count = mongo.db.users.count_documents({'club_id': club['_id']})
        c = serialize_doc(club)
        c['member_count'] = user_count
        result.append(c)
    return jsonify({'success': True, 'data': result})


# ============================================================
# SUPERADMIN: ENHANCED ENDPOINTS
# ============================================================

@api_bp.route('/superadmin/clubs/<club_id>/details', methods=['GET'])
@role_required('superadmin')
def superadmin_club_details(club_id):
    """Get detailed club info."""
    svc = get_platform_management_service()
    data = svc.get_club_details(club_id)
    if not data:
        return jsonify({'success': False, 'error': 'Club not found'}), 404
    data['club'] = serialize_doc(data['club'])
    data['teams'] = serialize_docs(data['teams'])
    return jsonify({'success': True, 'data': data})


@api_bp.route('/superadmin/clubs/<club_id>/suspend', methods=['POST'])
@role_required('superadmin')
def superadmin_suspend_club(club_id):
    """Suspend a club."""
    data = request.get_json() or {}
    svc = get_platform_management_service()
    svc.suspend_club(club_id, data.get('reason', ''))
    return jsonify({'success': True, 'message': 'Club suspendu'})


@api_bp.route('/superadmin/clubs/<club_id>/activate', methods=['POST'])
@role_required('superadmin')
def superadmin_activate_club(club_id):
    """Activate a club."""
    svc = get_platform_management_service()
    svc.activate_club(club_id)
    return jsonify({'success': True, 'message': 'Club activé'})


@api_bp.route('/superadmin/analytics', methods=['GET'])
@role_required('superadmin')
def superadmin_analytics():
    """Get platform analytics."""
    svc = get_platform_analytics_service()
    return jsonify({'success': True, 'data': svc.get_platform_metrics()})


@api_bp.route('/superadmin/analytics/growth', methods=['GET'])
@role_required('superadmin')
def superadmin_analytics_growth():
    """Get growth charts."""
    days = int(request.args.get('days', 90))
    svc = get_platform_analytics_service()
    return jsonify({'success': True, 'data': svc.get_growth_charts(days)})


@api_bp.route('/superadmin/analytics/revenue', methods=['GET'])
@role_required('superadmin')
def superadmin_analytics_revenue():
    """Get revenue breakdown by plan."""
    svc = get_platform_analytics_service()
    return jsonify({'success': True, 'data': svc.get_revenue_breakdown()})


@api_bp.route('/superadmin/analytics/cohorts', methods=['GET'])
@role_required('superadmin')
def superadmin_analytics_cohorts():
    """Get cohort analysis."""
    svc = get_platform_analytics_service()
    return jsonify({'success': True, 'data': svc.get_cohort_analysis()})


@api_bp.route('/superadmin/billing', methods=['GET'])
@role_required('superadmin')
def superadmin_billing():
    """Get platform billing overview."""
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_platform_billing()})


@api_bp.route('/superadmin/billing/subscriptions', methods=['GET'])
@role_required('superadmin')
def superadmin_billing_subscriptions():
    """Get all club subscriptions."""
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_all_subscriptions()})


@api_bp.route('/superadmin/billing/revenue', methods=['GET'])
@role_required('superadmin')
def superadmin_billing_revenue():
    """Get revenue chart data."""
    svc = get_billing_service()
    return jsonify({'success': True, 'data': svc.get_revenue_chart()})


@api_bp.route('/superadmin/billing/export/csv', methods=['GET'])
@role_required('superadmin')
def superadmin_billing_export_csv():
    """Export all subscriptions as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io

    svc = get_billing_service()
    subscriptions = svc.get_all_subscriptions()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Abonnements'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='22C55E')

    headers = ['Club', 'Plan', 'Statut', 'Montant (€/mois)', 'Membres', 'Équipes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for r, sub in enumerate(subscriptions, 2):
        ws.cell(row=r, column=1, value=sub.get('club_name', ''))
        ws.cell(row=r, column=2, value=sub.get('plan', ''))
        ws.cell(row=r, column=3, value=sub.get('status', ''))
        ws.cell(row=r, column=4, value=sub.get('amount', 0))
        ws.cell(row=r, column=5, value=sub.get('member_count', 0))
        ws.cell(row=r, column=6, value=sub.get('team_count', 0))

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=platform-billing.xlsx'
    return response


@api_bp.route('/superadmin/clubs/<club_id>/plan', methods=['PUT'])
@role_required('superadmin')
def superadmin_update_club_plan(club_id):
    """Update a club's subscription plan."""
    data = request.get_json()
    plan = data.get('plan') if data else None
    if not plan:
        return jsonify({'success': False, 'error': 'plan required'}), 400
    if ObjectId.is_valid(str(club_id)):
        mongo.db.clubs.update_one(
            {'_id': ObjectId(club_id)},
            {'$set': {'subscription.plan': plan, 'subscription.updated_at': datetime.datetime.utcnow().isoformat()}}
        )
    return jsonify({'success': True, 'message': 'Plan updated'})


@api_bp.route('/superadmin/analytics/export/pdf', methods=['GET'])
@role_required('superadmin')
def superadmin_analytics_export_pdf():
    """Export platform analytics as PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    svc = get_analytics_service()
    metrics = svc.get_platform_metrics() if hasattr(svc, 'get_platform_metrics') else {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('FootApp — Rapport Analytics Plateforme', styles['Title']))
    story.append(Spacer(1, 12))

    kpi_data = [['Indicateur', 'Valeur']]
    kpi_data.append(['Total clubs', str(metrics.get('total_clubs', 0))])
    kpi_data.append(['Total utilisateurs', str(metrics.get('total_users', 0))])
    kpi_data.append(['MAU', str(metrics.get('mau', 0))])
    kpi_data.append(['DAU', str(metrics.get('dau', 0))])
    kpi_data.append(['MRR (€)', str(metrics.get('mrr', 0))])
    kpi_data.append(['ARR (€)', str(metrics.get('arr', 0))])

    t = Table(kpi_data, colWidths=[250, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22c55e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f9fafb'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=platform-analytics.pdf'
    return response


@api_bp.route('/superadmin/analytics/export/excel', methods=['GET'])
@role_required('superadmin')
def superadmin_analytics_export_excel():
    """Export platform analytics as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io

    svc = get_analytics_service()
    metrics = svc.get_platform_metrics() if hasattr(svc, 'get_platform_metrics') else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KPIs Plateforme'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='22C55E')

    ws.cell(row=1, column=1, value='Indicateur').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value='Valeur').font = header_font
    ws.cell(row=1, column=2).fill = header_fill

    rows = [
        ('Total clubs', metrics.get('total_clubs', 0)),
        ('Total utilisateurs', metrics.get('total_users', 0)),
        ('MAU', metrics.get('mau', 0)),
        ('DAU', metrics.get('dau', 0)),
        ('MRR (€)', metrics.get('mrr', 0)),
        ('ARR (€)', metrics.get('arr', 0)),
    ]
    for r, (label, value) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=platform-analytics.xlsx'
    return response


# ============================================================
# SUPERADMIN: SUPPORT & MONITORING
# ============================================================

@api_bp.route('/superadmin/support/tickets', methods=['GET'])
@role_required('superadmin')
def superadmin_support_tickets():
    """List support tickets with optional filters."""
    status = request.args.get('status')
    priority = request.args.get('priority')
    club_id = request.args.get('club_id')

    query = {}
    if status:
        query['status'] = status
    if priority:
        query['priority'] = priority
    if club_id:
        query['club_id'] = club_id

    tickets = list(mongo.db.support_tickets.find(query).sort('created_at', -1).limit(200))
    return jsonify({'success': True, 'data': serialize_docs(tickets)})


@api_bp.route('/superadmin/support/tickets', methods=['POST'])
@role_required('superadmin')
def superadmin_create_support_ticket():
    """Create a new support ticket."""
    data = request.get_json()
    if not data or not data.get('title'):
        return jsonify({'success': False, 'error': 'title required'}), 400

    ticket = {
        'title': data.get('title', ''),
        'description': data.get('description', ''),
        'category': data.get('category', 'other'),
        'priority': data.get('priority', 'medium'),
        'status': 'open',
        'club_id': data.get('club_id', ''),
        'club_name': data.get('club_name', ''),
        'notes': [],
        'created_at': datetime.datetime.utcnow().isoformat(),
        'updated_at': datetime.datetime.utcnow().isoformat(),
    }
    result = mongo.db.support_tickets.insert_one(ticket)
    ticket['_id'] = result.inserted_id
    return jsonify({'success': True, 'data': serialize_doc(ticket)}), 201


@api_bp.route('/superadmin/support/tickets/<ticket_id>', methods=['PUT'])
@role_required('superadmin')
def superadmin_update_support_ticket(ticket_id):
    """Update a support ticket (status, priority, note)."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Data required'}), 400

    update = {'updated_at': datetime.datetime.utcnow().isoformat()}
    for field in ('status', 'priority', 'category', 'title', 'description'):
        if field in data:
            update[field] = data[field]

    ops = {'$set': update}
    if data.get('note'):
        ops['$push'] = {
            'notes': {
                'text': data['note'],
                'created_at': datetime.datetime.utcnow().isoformat(),
            }
        }

    if ObjectId.is_valid(str(ticket_id)):
        mongo.db.support_tickets.update_one({'_id': ObjectId(ticket_id)}, ops)
    return jsonify({'success': True, 'message': 'Ticket updated'})


@api_bp.route('/superadmin/support/monitoring', methods=['GET'])
@role_required('superadmin')
def superadmin_monitoring():
    """Get platform monitoring data: inactive clubs, low engagement clubs."""
    from datetime import datetime, timedelta

    cutoff_inactive = (datetime.utcnow() - timedelta(days=30)).isoformat()
    cutoff_veryinactive = (datetime.utcnow() - timedelta(days=60)).isoformat()

    clubs = list(mongo.db.clubs.find({}))
    inactive = []
    very_inactive = []

    for club in clubs:
        last_login = club.get('last_activity') or club.get('created_at', '')
        if last_login < cutoff_veryinactive:
            very_inactive.append({
                'id': str(club['_id']),
                'name': club.get('name', ''),
                'last_activity': last_login,
                'status': 'very_inactive',
            })
        elif last_login < cutoff_inactive:
            inactive.append({
                'id': str(club['_id']),
                'name': club.get('name', ''),
                'last_activity': last_login,
                'status': 'inactive',
            })

    return jsonify({
        'success': True,
        'data': {
            'inactive_clubs': inactive,
            'very_inactive_clubs': very_inactive,
            'total_monitored': len(clubs),
        }
    })


# ============================================================
# FAN ENDPOINTS
# ============================================================

@api_bp.route('/fan/comments/<post_id>', methods=['GET'])
@token_required
def fan_get_comments(post_id):
    """Get comments for a post."""
    svc = get_fan_engagement_service()
    return jsonify({'success': True, 'data': svc.get_comments(post_id)})


@api_bp.route('/fan/comments/<post_id>', methods=['POST'])
@token_required
def fan_create_comment(post_id):
    """Create a comment."""
    data = request.get_json()
    if not data or not data.get('content'):
        return jsonify({'success': False, 'error': 'Contenu requis'}), 400
    svc = get_fan_engagement_service()
    comment_id = svc.create_comment(post_id, request.current_user['user_id'],
                                     data['content'], data.get('parent_comment_id'))
    return jsonify({'success': True, 'comment_id': comment_id}), 201


@api_bp.route('/fan/reactions/<post_id>', methods=['POST'])
@token_required
def fan_toggle_reaction(post_id):
    """Toggle a reaction on a post."""
    data = request.get_json() or {}
    svc = get_fan_engagement_service()
    result = svc.toggle_reaction(post_id, request.current_user['user_id'], data.get('type', 'like'))
    return jsonify({'success': True, 'data': result})


@api_bp.route('/fan/polls', methods=['GET'])
@token_required
def fan_get_polls():
    """Get polls for club."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    svc = get_fan_engagement_service()
    return jsonify({'success': True, 'data': svc.get_polls(club_id)})


@api_bp.route('/fan/polls', methods=['POST'])
@role_required('admin', 'coach')
def fan_create_poll():
    """Create a poll (admin/coach only)."""
    data = request.get_json()
    if not data or not data.get('question') or not data.get('options'):
        return jsonify({'success': False, 'error': 'question et options requis'}), 400
    club_id = request.current_user.get('club_id')
    svc = get_fan_engagement_service()
    poll_id = svc.create_poll(club_id, data['question'], data['options'], data.get('expires_days', 7))
    return jsonify({'success': True, 'poll_id': poll_id}), 201


@api_bp.route('/fan/polls/<poll_id>/vote', methods=['POST'])
@token_required
def fan_vote_poll(poll_id):
    """Vote on a poll."""
    data = request.get_json()
    if data is None or 'option_index' not in data:
        return jsonify({'success': False, 'error': 'option_index requis'}), 400
    svc = get_fan_engagement_service()
    result = svc.vote_poll(poll_id, data['option_index'], request.current_user['user_id'])
    if result and result.get('error'):
        return jsonify({'success': False, 'error': result['error']}), 400
    return jsonify({'success': True})


@api_bp.route('/fan/media', methods=['GET'])
@token_required
def fan_get_media():
    """Get media gallery."""
    club_id = request.current_user.get('club_id')
    if not club_id:
        return jsonify({'success': True, 'data': []})
    category = request.args.get('category')
    svc = get_media_service()
    return jsonify({'success': True, 'data': svc.get_gallery(club_id, category)})


@api_bp.route('/fan/media/<media_id>', methods=['GET'])
@token_required
def fan_get_media_detail(media_id):
    """Get media detail."""
    svc = get_media_service()
    item = svc.get_media(media_id)
    if not item:
        return jsonify({'success': False, 'error': 'Media not found'}), 404
    return jsonify({'success': True, 'data': item})


@api_bp.route('/fan/media', methods=['POST'])
@role_required('admin', 'coach')
def fan_upload_media():
    """Upload media (admin/coach only)."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data'}), 400
    club_id = request.current_user.get('club_id')
    svc = get_media_service()
    media_id = svc.upload_media(club_id, data)
    return jsonify({'success': True, 'media_id': media_id}), 201


@api_bp.route('/matches/<match_id>/timeline', methods=['GET'])
def match_timeline(match_id):
    """Get match timeline (public)."""
    match = mongo.db.matches.find_one({'_id': ObjectId(match_id)})
    if not match:
        return jsonify({'success': False, 'error': 'Match not found'}), 404
    events = match.get('events', [])
    return jsonify({'success': True, 'data': events})


@api_bp.route('/matches/<match_id>/stats', methods=['GET'])
def match_stats(match_id):
    """Get match statistics (public)."""
    match = mongo.db.matches.find_one({'_id': ObjectId(match_id)})
    if not match:
        return jsonify({'success': False, 'error': 'Match not found'}), 404
    stats = match.get('stats', {
        'possession_home': 50, 'possession_away': 50,
        'shots_home': 0, 'shots_away': 0,
        'corners_home': 0, 'corners_away': 0,
    })
    return jsonify({'success': True, 'data': stats})


@api_bp.route('/matches/fixtures/<club_id>', methods=['GET'])
def match_fixtures(club_id):
    """Get upcoming fixtures (public)."""
    matches = list(mongo.db.matches.find({
        'club_id': ObjectId(club_id),
        'status': {'$ne': 'completed'}
    }).sort('date', 1).limit(10))
    return jsonify({'success': True, 'data': serialize_docs(matches)})
