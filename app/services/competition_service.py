# FootLogic V2 - Competition Service

from bson import ObjectId
from datetime import datetime


class CompetitionService:
    """Service for competition/tournament management"""

    def __init__(self, db):
        self.db = db
        self.collection = db.competitions

    # ── CRUD ──────────────────────────────────────────────

    def get_all(self, club_id):
        """Get all competitions for a club, newest first"""
        return list(self.collection.find({'club_id': ObjectId(club_id)}).sort('created_at', -1))

    def get_by_id(self, competition_id):
        return self.collection.find_one({'_id': ObjectId(competition_id)})

    def create(self, club_id, data):
        doc = {
            'club_id': ObjectId(club_id),
            'name': data['name'],
            'type': data.get('type', 'league'),          # league / cup / tournament / friendly
            'season': data.get('season', ''),
            'category': data.get('category', ''),         # Senior, U19, U17 …
            'organizer': data.get('organizer', ''),
            'start_date': data.get('start_date'),
            'end_date': data.get('end_date'),
            'status': data.get('status', 'active'),       # active / finished / upcoming
            'teams_involved': data.get('teams_involved', []),  # list of team_id strings
            'notes': data.get('notes', ''),
            'created_at': datetime.utcnow(),
        }
        result = self.collection.insert_one(doc)
        doc['_id'] = result.inserted_id
        return doc

    def update(self, competition_id, data):
        allowed = ['name', 'type', 'season', 'category', 'organizer',
                    'start_date', 'end_date', 'status', 'teams_involved', 'notes']
        update_fields = {k: v for k, v in data.items() if k in allowed}
        if update_fields:
            self.collection.update_one(
                {'_id': ObjectId(competition_id)},
                {'$set': update_fields}
            )

    def delete(self, competition_id):
        return self.collection.delete_one({'_id': ObjectId(competition_id)})

    # ── Excel helpers ─────────────────────────────────────

    def export_excel(self, club_id):
        """Export competitions to Excel bytes (openpyxl)"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        comps = self.get_all(club_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compétitions"

        header_fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        headers = ['Nom', 'Type', 'Saison', 'Catégorie', 'Organisateur',
                   'Date début', 'Date fin', 'Statut', 'Notes']
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        type_map = {'league': 'Championnat', 'cup': 'Coupe', 'tournament': 'Tournoi', 'friendly': 'Amical'}
        status_map = {'active': 'En cours', 'finished': 'Terminé', 'upcoming': 'À venir'}

        for c in comps:
            ws.append([
                c.get('name', ''),
                type_map.get(c.get('type', ''), c.get('type', '')),
                c.get('season', ''),
                c.get('category', ''),
                c.get('organizer', ''),
                c['start_date'].strftime('%d/%m/%Y') if isinstance(c.get('start_date'), datetime) else str(c.get('start_date', '')),
                c['end_date'].strftime('%d/%m/%Y') if isinstance(c.get('end_date'), datetime) else str(c.get('end_date', '')),
                status_map.get(c.get('status', ''), c.get('status', '')),
                c.get('notes', ''),
            ])

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col_letter].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def import_excel(self, club_id, file_bytes):
        """Import competitions from Excel. Returns (created_count, errors)."""
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        errors = []

        type_rev = {'championnat': 'league', 'coupe': 'cup', 'tournoi': 'tournament', 'amical': 'friendly',
                    'league': 'league', 'cup': 'cup', 'tournament': 'tournament', 'friendly': 'friendly'}
        status_rev = {'en cours': 'active', 'terminé': 'finished', 'à venir': 'upcoming',
                      'active': 'active', 'finished': 'finished', 'upcoming': 'upcoming'}

        for idx, row in enumerate(rows, start=2):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name:
                errors.append(f"Ligne {idx}: nom manquant")
                continue

            comp_type = type_rev.get(str(row[1] or '').strip().lower(), 'league')
            season = str(row[2] or '').strip()
            category = str(row[3] or '').strip()
            organizer = str(row[4] or '').strip()

            start_date = self._parse_date(row[5])
            end_date = self._parse_date(row[6])
            status = status_rev.get(str(row[7] or '').strip().lower(), 'active')
            notes = str(row[8] or '').strip() if len(row) > 8 else ''

            self.create(club_id, {
                'name': name, 'type': comp_type, 'season': season,
                'category': category, 'organizer': organizer,
                'start_date': start_date, 'end_date': end_date,
                'status': status, 'notes': notes,
            })
            created += 1

        return created, errors

    @staticmethod
    def _parse_date(val):
        if isinstance(val, datetime):
            return val
        if not val:
            return None
        val_str = str(val).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
        return None
