# FootLogic V2 - Match Service

from bson import ObjectId
from datetime import datetime

class MatchService:
    """Service for match-related operations"""

    def __init__(self, db):
        self.db = db
        self.collection = db.matches

    def get_all(self):
        """Get all matches sorted by date"""
        return list(self.collection.find().sort('date', -1))

    def get_by_id(self, match_id):
        """Get match by ID"""
        return self.collection.find_one({'_id': ObjectId(match_id)})

    def get_by_club(self, club_id):
        """Get all matches for a club"""
        return list(self.collection.find({'club_id': ObjectId(club_id)}).sort('date', -1))

    def get_upcoming(self, club_id, team_id=None, limit=5):
        """Get upcoming scheduled matches, optionally filtered by team"""
        query = {
            'club_id': ObjectId(club_id),
            'status': 'scheduled'
        }
        if team_id:
            query['team_id'] = ObjectId(team_id)
        return list(self.collection.find(query).sort('date', 1).limit(limit))

    def get_completed(self, club_id, team_id=None, limit=10):
        """Get completed matches, optionally filtered by team"""
        query = {
            'club_id': ObjectId(club_id),
            'status': 'completed'
        }
        if team_id:
            query['team_id'] = ObjectId(team_id)
        return list(self.collection.find(query).sort('date', -1).limit(limit))

    def create(self, club_id, opponent, date, is_home=True, **kwargs):
        """Create a new match"""
        match = {
            'club_id': ObjectId(club_id),
            'opponent': opponent,
            'date': date,
            'location': kwargs.get('location', ''),
            'is_home': is_home,
            'score': {'home': 0, 'away': 0},
            'status': kwargs.get('status', 'scheduled'),
            'lineup': [],
            'events': [],
            'created_at': datetime.utcnow()
        }
        result = self.collection.insert_one(match)
        match['_id'] = result.inserted_id
        return match

    def update(self, match_id, data):
        """Update match data"""
        return self.collection.update_one(
            {'_id': ObjectId(match_id)},
            {'$set': data}
        )

    def set_score(self, match_id, home_score, away_score, status=None):
        """Update match score and optionally status"""
        update_data = {
            'score.home': home_score,
            'score.away': away_score
        }
        if status:
            update_data['status'] = status

        return self.collection.update_one(
            {'_id': ObjectId(match_id)},
            {'$set': update_data}
        )

    def start_match(self, match_id):
        """Set match status to live"""
        return self.collection.update_one(
            {'_id': ObjectId(match_id)},
            {'$set': {'status': 'live'}}
        )

    def finish_match(self, match_id):
        """Set match status to completed"""
        return self.collection.update_one(
            {'_id': ObjectId(match_id)},
            {'$set': {'status': 'completed'}}
        )

    def set_lineup(self, match_id, player_ids):
        """Set match lineup"""
        return self.collection.update_one(
            {'_id': ObjectId(match_id)},
            {'$set': {'lineup': [ObjectId(pid) for pid in player_ids]}}
        )

    def add_event(self, match_id, event_type, player_id, minute):
        """Add match event (goal, assist, card, substitution)"""
        event = {
            'type': event_type,
            'player_id': ObjectId(player_id),
            'minute': minute,
            'timestamp': datetime.utcnow()
        }
        return self.collection.update_one(
            {'_id': ObjectId(match_id)},
            {'$push': {'events': event}}
        )

    def get_lineup(self, match_id):
        """Get match lineup with player details"""
        match = self.get_by_id(match_id)
        if not match:
            return []

        lineup_ids = match.get('lineup', [])
        return list(self.db.players.find({'_id': {'$in': lineup_ids}}))

    def get_season_stats(self, club_id, team_id=None):
        """Get season statistics for a club, optionally filtered by team"""
        query = {
            'club_id': ObjectId(club_id),
            'status': 'completed'
        }
        if team_id:
            query['team_id'] = ObjectId(team_id)

        matches = list(self.collection.find(query))

        wins = draws = losses = goals_for = goals_against = 0

        for m in matches:
            score = m.get('score', {'home': 0, 'away': 0})
            is_home = m.get('is_home', True)

            our_goals = score['home'] if is_home else score['away']
            their_goals = score['away'] if is_home else score['home']

            goals_for += our_goals
            goals_against += their_goals

            if our_goals > their_goals:
                wins += 1
            elif our_goals < their_goals:
                losses += 1
            else:
                draws += 1

        return {
            'played': len(matches),
            'wins': wins,
            'draws': draws,
            'losses': losses,
            'goals_for': goals_for,
            'goals_against': goals_against,
            'goal_difference': goals_for - goals_against,
            'points': (wins * 3) + draws
        }

    def delete(self, match_id):
        """Delete a match"""
        return self.collection.delete_one({'_id': ObjectId(match_id)})

    # ── Excel helpers ─────────────────────────────────────

    def export_excel(self, club_id):
        """Export matches to Excel bytes"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io

        matches = self.get_by_club(club_id)
        # Resolve team names
        team_map = {}
        team_ids = set()
        for m in matches:
            tid = m.get('team_id')
            if tid:
                team_ids.add(tid)
        if team_ids:
            teams = list(self.db.teams.find({'_id': {'$in': list(team_ids)}}))
            team_map = {str(t['_id']): t.get('name', '') for t in teams}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matchs"

        hfill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)

        headers = ['Adversaire', 'Date', 'Heure', 'Lieu', 'Domicile/Extérieur',
                   'Équipe', 'Statut', 'Score Domicile', 'Score Extérieur', 'Compétition']
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center')

        status_map = {'scheduled': 'Programmé', 'live': 'En cours', 'completed': 'Terminé', 'cancelled': 'Annulé'}

        for m in matches:
            d = m.get('date')
            score = m.get('score', {})
            ws.append([
                m.get('opponent', ''),
                d.strftime('%d/%m/%Y') if isinstance(d, datetime) else str(d or ''),
                d.strftime('%H:%M') if isinstance(d, datetime) else '',
                m.get('location', ''),
                'Domicile' if m.get('is_home') else 'Extérieur',
                team_map.get(str(m.get('team_id', '')), ''),
                status_map.get(m.get('status', ''), m.get('status', '')),
                score.get('home', 0),
                score.get('away', 0),
                m.get('competition', ''),
            ])

        for i, w in enumerate([22, 14, 10, 22, 18, 18, 14, 16, 16, 22], start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def import_excel(self, club_id, file_bytes, team_map=None):
        """Import matches from Excel. Returns (created_count, errors).
        team_map: dict {team_name_lower: team_id_str}
        """
        import openpyxl
        import io as _io

        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        errors = []

        home_vals = {'domicile', 'dom', 'home', 'oui', 'yes', '1', 'true'}

        for idx, row in enumerate(rows, start=2):
            if not row or not row[0]:
                continue
            opponent = str(row[0]).strip()
            if not opponent:
                errors.append(f"Ligne {idx}: adversaire manquant")
                continue

            # Date
            date_val = row[1]
            match_date = self._parse_datetime(date_val, row[2] if len(row) > 2 else None)
            if not match_date:
                errors.append(f"Ligne {idx}: date invalide")
                continue

            location = str(row[3] or '').strip() if len(row) > 3 else ''
            is_home = str(row[4] or '').strip().lower() in home_vals if len(row) > 4 else True
            team_name = str(row[5] or '').strip().lower() if len(row) > 5 else ''
            status_raw = str(row[6] or '').strip().lower() if len(row) > 6 else 'scheduled'

            status_rev = {'programmé': 'scheduled', 'en cours': 'live', 'terminé': 'completed',
                          'annulé': 'cancelled', 'scheduled': 'scheduled', 'live': 'live',
                          'completed': 'completed', 'cancelled': 'cancelled'}
            status = status_rev.get(status_raw, 'scheduled')

            score_home = self._safe_int(row[7]) if len(row) > 7 else 0
            score_away = self._safe_int(row[8]) if len(row) > 8 else 0
            competition = str(row[9] or '').strip() if len(row) > 9 else ''

            kwargs = {
                'location': location,
                'status': status,
            }

            match = self.create(club_id, opponent, match_date, is_home=is_home, **kwargs)

            # Set team_id if resolved
            if team_map and team_name and team_name in team_map:
                self.collection.update_one(
                    {'_id': match['_id']},
                    {'$set': {'team_id': ObjectId(team_map[team_name])}}
                )

            # Set score & competition
            updates = {}
            if score_home or score_away:
                updates['score'] = {'home': score_home, 'away': score_away}
            if competition:
                updates['competition'] = competition
            if updates:
                self.collection.update_one({'_id': match['_id']}, {'$set': updates})

            created += 1

        return created, errors

    @staticmethod
    def _parse_datetime(date_val, time_val=None):
        if isinstance(date_val, datetime):
            return date_val
        if not date_val:
            return None
        date_str = str(date_val).strip()
        parsed = None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                parsed = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        if not parsed:
            return None
        if time_val:
            time_str = str(time_val).strip()
            for tfmt in ('%H:%M', '%H:%M:%S', '%Hh%M'):
                try:
                    t = datetime.strptime(time_str, tfmt)
                    parsed = parsed.replace(hour=t.hour, minute=t.minute)
                    break
                except ValueError:
                    continue
        return parsed

    @staticmethod
    def _safe_int(val):
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return 0
