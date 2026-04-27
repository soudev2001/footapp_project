# FootLogic V2 - Analytics Service
# Provides club-level analytics for the admin dashboard

from datetime import datetime, timedelta
from bson import ObjectId
import io


class AnalyticsService:
    def __init__(self, db):
        self.db = db
        self.users = db['users']
        self.teams = db['teams']
        self.players = db['players']
        self.events = db['events']
        self.messages = db['messages']

    # ================================================================
    # MEMBER METRICS
    # ================================================================

    def get_member_growth(self, club_id, days=90):
        """Return daily new member counts over the last N days."""
        since = datetime.utcnow() - timedelta(days=days)
        pipeline = [
            {'$match': {
                'club_id': ObjectId(club_id),
                'created_at': {'$gte': since}
            }},
            {'$group': {
                '_id': {
                    'y': {'$year': '$created_at'},
                    'm': {'$month': '$created_at'},
                    'd': {'$dayOfMonth': '$created_at'}
                },
                'count': {'$sum': 1}
            }},
            {'$sort': {'_id.y': 1, '_id.m': 1, '_id.d': 1}}
        ]
        raw = list(self.users.aggregate(pipeline))
        labels, data = [], []
        for r in raw:
            d = r['_id']
            labels.append(f"{d['d']:02d}/{d['m']:02d}")
            data.append(r['count'])
        return {'labels': labels, 'data': data}

    def get_members_by_role(self, club_id):
        """Return member count grouped by role."""
        pipeline = [
            {'$match': {'club_id': ObjectId(club_id)}},
            {'$group': {'_id': '$role', 'count': {'$sum': 1}}}
        ]
        result = {r['_id']: r['count'] for r in self.users.aggregate(pipeline)}
        roles = ['admin', 'coach', 'player', 'parent', 'fan']
        return {
            'labels': roles,
            'data': [result.get(r, 0) for r in roles]
        }

    def get_engagement_metrics(self, club_id):
        """Return active vs inactive members (active = logged in within 30 days)."""
        threshold = datetime.utcnow() - timedelta(days=30)
        total = self.users.count_documents({'club_id': ObjectId(club_id)})
        active = self.users.count_documents({
            'club_id': ObjectId(club_id),
            'last_login': {'$gte': threshold}
        })
        inactive = total - active

        # Top active members (most logins)
        top_active = list(self.users.find(
            {'club_id': ObjectId(club_id), 'login_count': {'$gt': 0}},
            {'profile': 1, 'role': 1, 'login_count': 1, 'last_login': 1}
        ).sort('login_count', -1).limit(5))

        return {
            'total': total,
            'active': active,
            'inactive': inactive,
            'active_pct': round((active / total * 100) if total else 0, 1),
            'top_active': top_active
        }

    def get_new_members_this_month(self, club_id):
        """Return count of members created this calendar month."""
        now = datetime.utcnow()
        start_of_month = datetime(now.year, now.month, 1)
        return self.users.count_documents({
            'club_id': ObjectId(club_id),
            'created_at': {'$gte': start_of_month}
        })

    # ================================================================
    # TEAM METRICS
    # ================================================================

    def get_team_stats(self, club_id):
        """Return player and coach count per team."""
        teams = list(self.teams.find({'club_id': ObjectId(club_id)}))
        stats = []
        for team in teams:
            player_count = self.players.count_documents({'team_id': team['_id']})
            coach_count = len(team.get('coach_ids', []))
            stats.append({
                'name': team['name'],
                'category': team.get('category', ''),
                'players': player_count,
                'coaches': coach_count,
                'colors': team.get('colors', {})
            })
        labels = [t['name'] for t in stats]
        player_data = [t['players'] for t in stats]
        coach_data = [t['coaches'] for t in stats]
        return {
            'labels': labels,
            'player_data': player_data,
            'coach_data': coach_data,
            'details': stats
        }

    def get_team_performance(self, club_id):
        """Return win/draw/loss and attendance per team."""
        teams = list(self.teams.find({'club_id': ObjectId(club_id)}))
        result = []
        for team in teams:
            tid = team['_id']
            matches = list(self.db['matches'].find({'$or': [
                {'club_id': ObjectId(club_id), 'team_id': tid},
                {'club_id': ObjectId(club_id)}
            ], 'status': 'completed'}))
            w = d = l = 0
            for m in matches:
                h = m.get('home_score', 0)
                a = m.get('away_score', 0)
                if h > a:
                    w += 1
                elif h == a:
                    d += 1
                else:
                    l += 1
            events = list(self.events.find({'club_id': ObjectId(club_id), 'team_id': tid, 'type': 'training'}))
            total_att = 0
            present_att = 0
            for ev in events:
                for att in ev.get('attendance', []):
                    total_att += 1
                    if att.get('status') == 'present':
                        present_att += 1
            att_rate = round((present_att / total_att * 100) if total_att else 0, 1)
            result.append({
                'team_name': team['name'],
                'category': team.get('category', ''),
                'wins': w, 'draws': d, 'losses': l,
                'matches_total': w + d + l,
                'attendance_rate': att_rate,
            })
        return result

    def get_member_retention(self, club_id, days=90):
        """Return retention and churn metrics."""
        now = datetime.utcnow()
        total = self.users.count_documents({'club_id': ObjectId(club_id)})
        period_start = now - timedelta(days=days)
        new_3m = self.users.count_documents({
            'club_id': ObjectId(club_id),
            'created_at': {'$gte': period_start}
        })
        still_active_3m = self.users.count_documents({
            'club_id': ObjectId(club_id),
            'created_at': {'$gte': period_start},
            'last_login': {'$gte': now - timedelta(days=30)}
        })
        retention_3m = round((still_active_3m / new_3m * 100) if new_3m else 0, 1)
        one_month_ago = now - timedelta(days=30)
        left_last_month = self.users.count_documents({
            'club_id': ObjectId(club_id),
            'account_status': {'$in': ['deleted', 'suspended']},
            'updated_at': {'$gte': one_month_ago}
        })
        churn_rate = round((left_last_month / total * 100) if total else 0, 1)
        return {
            'total': total,
            'new_3_months': new_3m,
            'retained_3_months': still_active_3m,
            'retention_rate_3m': retention_3m,
            'churn_rate_monthly': churn_rate,
        }

    def get_feature_usage(self, club_id):
        """Return basic feature usage stats."""
        cid = ObjectId(club_id)
        return {
            'messages_sent': self.messages.count_documents({'club_id': cid}),
            'events_created': self.events.count_documents({'club_id': cid}),
            'matches_played': self.db['matches'].count_documents({'club_id': cid, 'status': 'completed'}),
            'posts_created': self.db['posts'].count_documents({'club_id': cid}),
            'players_registered': self.players.count_documents({'club_id': cid}),
        }

    # ================================================================
    # FINANCIAL METRICS
    # ================================================================

    def get_financial_metrics(self, club_id):
        """Return MRR and subscription info from clubs collection."""
        club = self.db['clubs'].find_one({'_id': ObjectId(club_id)})
        if not club:
            return {'mrr': 0, 'arr': 0, 'plan': 'Aucun', 'status': 'N/A'}

        subscription = club.get('subscription', {})
        billing = subscription.get('billing', {})
        mrr = float(billing.get('total_monthly', 0))
        return {
            'mrr': mrr,
            'arr': round(mrr * 12, 2),
            'plan': subscription.get('plan_id', 'Gratuit').replace('_', ' ').title(),
            'status': subscription.get('status', 'N/A'),
            'revenue_per_member': 0  # Extended in Phase 4 with Stripe
        }

    # ================================================================
    # FULL DASHBOARD SUMMARY
    # ================================================================

    def get_dashboard_summary(self, club_id):
        """Aggregate all metrics in one call."""
        engagement = self.get_engagement_metrics(club_id)
        return {
            'member_growth': self.get_member_growth(club_id, days=90),
            'members_by_role': self.get_members_by_role(club_id),
            'team_stats': self.get_team_stats(club_id),
            'engagement': engagement,
            'new_this_month': self.get_new_members_this_month(club_id),
            'financial': self.get_financial_metrics(club_id)
        }

    # ================================================================
    # EXPORT: PDF
    # ================================================================

    def export_pdf(self, club_id):
        """Generate a PDF report and return bytes."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.units import cm
        except ImportError:
            return None

        summary = self.get_dashboard_summary(club_id)
        club = self.db['clubs'].find_one({'_id': ObjectId(club_id)})
        club_name = club.get('name', 'Club') if club else 'Club'
        now = datetime.utcnow()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=2 * cm, leftMargin=2 * cm,
                                topMargin=2 * cm, bottomMargin=2 * cm)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle('title', parent=styles['Title'],
                                     fontSize=22, textColor=colors.HexColor('#1e40af'))
        story.append(Paragraph(f"Rapport Club — {club_name}", title_style))
        story.append(Paragraph(f"Généré le {now.strftime('%d/%m/%Y à %H:%M')} UTC", styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        # --- Member section ---
        story.append(Paragraph("Membres", styles['Heading2']))
        eng = summary['engagement']
        member_data = [
            ['Indicateur', 'Valeur'],
            ['Total membres', str(eng['total'])],
            ['Actifs (30 derniers jours)', f"{eng['active']} ({eng['active_pct']}%)"],
            ['Inactifs', str(eng['inactive'])],
            ['Nouveaux ce mois-ci', str(summary['new_this_month'])],
        ]
        for role, count in zip(summary['members_by_role']['labels'],
                               summary['members_by_role']['data']):
            if count:
                member_data.append([f'  Rôle: {role.capitalize()}', str(count)])

        member_table = Table(member_data, colWidths=[10 * cm, 6 * cm])
        member_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(member_table)
        story.append(Spacer(1, 0.5 * cm))

        # --- Teams section ---
        story.append(Paragraph("Équipes", styles['Heading2']))
        team_stats = summary['team_stats']['details']
        if team_stats:
            team_data = [['Équipe', 'Catégorie', 'Joueurs', 'Entraîneurs']]
            for t in team_stats:
                team_data.append([t['name'], t.get('category', ''), str(t['players']), str(t['coaches'])])
            team_table = Table(team_data, colWidths=[6 * cm, 4 * cm, 4 * cm, 4 * cm])
            team_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(team_table)
        else:
            story.append(Paragraph("Aucune équipe enregistrée.", styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        # --- Financial section ---
        story.append(Paragraph("Finances", styles['Heading2']))
        fin = summary['financial']
        fin_data = [
            ['Indicateur', 'Valeur'],
            ['Plan actif', fin['plan']],
            ['Statut', fin['status']],
            ['MRR', f"{fin['mrr']} €"],
            ['ARR (projection)', f"{fin['arr']} €"],
        ]
        fin_table = Table(fin_data, colWidths=[10 * cm, 6 * cm])
        fin_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#faf5ff')]),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(fin_table)

        doc.build(story)
        return buf.getvalue()

    # ================================================================
    # EXPORT: Excel
    # ================================================================

    def export_excel(self, club_id):
        """Generate an Excel report and return bytes."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return None

        summary = self.get_dashboard_summary(club_id)
        club = self.db['clubs'].find_one({'_id': ObjectId(club_id)})
        club_name = club.get('name', 'Club') if club else 'Club'
        now = datetime.utcnow()

        wb = openpyxl.Workbook()

        # --- Sheet 1: Membres ---
        ws = wb.active
        ws.title = "Membres"
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

        ws['A1'] = f"Rapport Club — {club_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Généré le {now.strftime('%d/%m/%Y à %H:%M')} UTC"
        ws.append([])

        ws.append(['Indicateur', 'Valeur'])
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font

        eng = summary['engagement']
        rows = [
            ('Total membres', eng['total']),
            ('Actifs (30j)', eng['active']),
            ('Inactifs', eng['inactive']),
            ('% actifs', f"{eng['active_pct']}%"),
            ('Nouveaux ce mois', summary['new_this_month']),
        ]
        for i, (k, v) in enumerate(rows):
            ws.append([k, v])
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = alt_fill
        ws.append([])

        ws.append(['Rôle', 'Nombre'])
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        for role, count in zip(summary['members_by_role']['labels'],
                               summary['members_by_role']['data']):
            ws.append([role.capitalize(), count])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        # --- Sheet 2: Équipes ---
        ws2 = wb.create_sheet("Équipes")
        ws2.append(['Équipe', 'Catégorie', 'Joueurs', 'Entraîneurs'])
        for cell in ws2[1]:
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            cell.font = header_font
        for t in summary['team_stats']['details']:
            ws2.append([t['name'], t.get('category', ''), t['players'], t['coaches']])
        for col in ['A', 'B', 'C', 'D']:
            ws2.column_dimensions[col].width = 20

        # --- Sheet 3: Finances ---
        ws3 = wb.create_sheet("Finances")
        fin = summary['financial']
        ws3.append(['Indicateur', 'Valeur'])
        for cell in ws3[1]:
            cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            cell.font = header_font
        ws3.append(['Plan actif', fin['plan']])
        ws3.append(['Statut', fin['status']])
        ws3.append(['MRR', f"{fin['mrr']} €"])
        ws3.append(['ARR (projection)', f"{fin['arr']} €"])
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
